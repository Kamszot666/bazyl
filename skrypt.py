#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SKRYPT UZUPELNIANIA BAZY KONTAKTOW FIRM
Wersja: 3.0
Opis: Czyta Excel -> scrapuje FB/LinkedIn/Email/Telefon/NIP osoby kontaktowej -> zapisuje wyniki
Kompatybilnosc: Windows 11 Pro + NVDA + Selenium 4.21+ + Python 3.10+
"""

# =============================================================================
# KONFIGURACJA - ZMIEN PONIZSZE WARTOSCI WEDLUG POTRZEB
# =============================================================================

# --- PLIK EXCEL ---
NAZWA_ARKUSZA = "baza.xlsx"           # <- ZMIEN nazwe swojego pliku Excel
SCIEZKA_WYNIKI = "wyniki/"            # <- Folder wynikow (tworzony automatycznie)

# --- KOLUMNY W EXCELU (numery kolumn, A=1, B=2, ...) ---
KOL_NAZWA_FIRMY = 2                  # <- Kolumna z nazwa firmy (B=2)
KOL_WWW = 6                          # <- Kolumna ze strona www (F=6)
KOL_FACEBOOK = 8                     # <- Kolumna na Facebook (H=8)
KOL_LINKEDIN = 9                     # <- Kolumna na LinkedIn (I=9)
KOL_EMAIL = 10                       # <- Kolumna na Email (J=10)
KOL_TELEFON = 11                     # <- Kolumna na Telefon (K=11)
KOL_NIP = 12                         # <- Kolumna na NIP (L=12)
KOL_IMIE_NAZWISKO = 13               # <- Kolumna na imie i nazwisko (M=13)
KOL_STANOWISKO = 14                  # <- Kolumna na stanowisko (N=14)

# --- PARAMETRY SCRAPOWANIA ---
MAX_WATKOW = 4                        # <- ZMIEN 1-8 watkow rownoczesnych
TIMEOUT_STRONA = 15                   # <- ZMIEN sekundy timeout na strone
RETRIES = 3                           # <- ZMIEN liczbe ponownych prob
OPOZNIENIE_MIN = 1.0                  # <- Minimum sekund miedzy zapytaniami
OPOZNIENIE_MAX = 2.0                  # <- Maksimum sekund miedzy zapytaniami
MAX_WYSZUKIWAN_DDG = 5                # <- DuckDuckGo max zapytan na firme
CHECKPOINT_CO = 10                    # <- Zapis stanu co N rekordow
HEADLESS = True                       # <- False = widoczna przegladarka, True = ukryta
WIERSZ_STARTOWY = 2                   # <- Od ktorego wiersza zaczac (2 = pierwszy po naglowku)

# =============================================================================
# KONIEC KONFIGURACJI - PONIZEJ NIE TRZEBA NIC ZMIENIAC
# =============================================================================

import sys
import os
import re
import csv
import json
import time
import logging
import threading
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin

# Sprawdzenie i import bibliotek
try:
    import openpyxl
except ImportError:
    print("BLAD: Brak biblioteki openpyxl. Zainstaluj: pip install openpyxl")
    sys.exit(1)

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("BLAD: Brak biblioteki beautifulsoup4. Zainstaluj: pip install beautifulsoup4")
    sys.exit(1)

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import (
        TimeoutException,
        NoSuchElementException,
        WebDriverException,
    )
except ImportError:
    print("BLAD: Brak biblioteki selenium. Zainstaluj: pip install selenium==4.21.0")
    sys.exit(1)

try:
    from webdriver_manager.chrome import ChromeDriverManager
except ImportError:
    print("BLAD: Brak biblioteki webdriver-manager. Zainstaluj: pip install webdriver-manager")
    sys.exit(1)


# =============================================================================
# SCIEZKI PLIKOW
# =============================================================================
FOLDER_SKRYPTU = Path(__file__).parent.resolve()
SCIEZKA_EXCEL = FOLDER_SKRYPTU / NAZWA_ARKUSZA
SCIEZKA_WYNIKOW = FOLDER_SKRYPTU / SCIEZKA_WYNIKI
SCIEZKA_LOGU = FOLDER_SKRYPTU / "logi.csv"
SCIEZKA_CHECKPOINT = FOLDER_SKRYPTU / "checkpoint.json"
SCIEZKA_CONFIG = FOLDER_SKRYPTU / "config.json"
SCIEZKA_BAT = FOLDER_SKRYPTU / "skrypt.bat"


# =============================================================================
# KONFIGURACJA LOGOWANIA
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(FOLDER_SKRYPTU / "scraping.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("skrypt")


# =============================================================================
# POMOCNICZE FUNKCJE NVDA
# =============================================================================

def mow(tekst: str) -> None:
    """Wyswietla tekst w sposob przyjazny dla NVDA (bez kolorow ANSI)."""
    print(f"\n{tekst}\n")
    time.sleep(0.5)


def mow_postep(numer: int, total: int, nazwa: str) -> None:
    """Wyswietla postep przetwarzania w sposob przyjazny dla NVDA."""
    procent = int((numer / total) * 100) if total > 0 else 0
    print(f"\n{numer}/{total} firm ({procent}%) - {nazwa}\n")


def zapytaj(pytanie: str) -> str:
    """Zadaje pytanie uzytkownikowi w sposob przyjazny dla NVDA."""
    time.sleep(0.5)
    odpowiedz = input(f"\n{pytanie}\n")
    time.sleep(0.5)
    return odpowiedz.strip()


# =============================================================================
# LOGOWANIE DO CSV
# =============================================================================

def zapisz_log_csv(zdarzenie: str) -> None:
    """Zapisuje zdarzenie do pliku logi.csv."""
    try:
        plik_istnieje = SCIEZKA_LOGU.exists()
        with open(SCIEZKA_LOGU, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not plik_istnieje:
                writer.writerow(["zdarzenie", "czas"])
            writer.writerow([zdarzenie, datetime.now().isoformat()])
    except Exception as e:
        logger.error(f"Blad zapisu logu CSV: {e}")


# =============================================================================
# CHECKPOINT
# =============================================================================

def zapisz_checkpoint(ostatni_wiersz: int) -> None:
    """Zapisuje checkpoint - numer ostatniego przetworzonego wiersza."""
    try:
        dane = {"ostatni_wiersz": ostatni_wiersz, "czas": datetime.now().isoformat()}
        with open(SCIEZKA_CHECKPOINT, "w", encoding="utf-8") as f:
            json.dump(dane, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"Blad zapisu checkpointu: {e}")


def wczytaj_checkpoint() -> Optional[int]:
    """Wczytuje checkpoint - numer ostatniego przetworzonego wiersza."""
    try:
        if SCIEZKA_CHECKPOINT.exists():
            with open(SCIEZKA_CHECKPOINT, "r", encoding="utf-8") as f:
                dane = json.load(f)
            return dane.get("ostatni_wiersz")
    except Exception as e:
        logger.warning(f"Blad wczytywania checkpointu: {e}")
    return None


# =============================================================================
# GENEROWANIE PLIKOW KONFIGURACYJNYCH
# =============================================================================

def generuj_config_json() -> None:
    """Generuje plik config.json z aktualna konfiguracja."""
    config = {
        "nazwa_arkusza": NAZWA_ARKUSZA,
        "sciezka_wyniki": SCIEZKA_WYNIKI,
        "max_watkow": MAX_WATKOW,
        "timeout_strona": TIMEOUT_STRONA,
        "retries": RETRIES,
        "opoznienie_min": OPOZNIENIE_MIN,
        "opoznienie_max": OPOZNIENIE_MAX,
        "max_wyszukiwan_ddg": MAX_WYSZUKIWAN_DDG,
        "checkpoint_co": CHECKPOINT_CO,
        "headless": HEADLESS,
        "kolumny": {
            "nazwa_firmy": KOL_NAZWA_FIRMY,
            "www": KOL_WWW,
            "facebook": KOL_FACEBOOK,
            "linkedin": KOL_LINKEDIN,
            "email": KOL_EMAIL,
            "telefon": KOL_TELEFON,
            "nip": KOL_NIP,
            "imie_nazwisko": KOL_IMIE_NAZWISKO,
            "stanowisko": KOL_STANOWISKO,
        },
    }
    try:
        with open(SCIEZKA_CONFIG, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        logger.info(f"Wygenerowano config.json")
    except Exception as e:
        logger.error(f"Blad generowania config.json: {e}")


def generuj_bat() -> None:
    """Generuje plik skrypt.bat do uruchamiania skryptu."""
    tresc = (
        "@echo off\r\n"
        'cd /d "%~dp0"\r\n'
        "echo Uruchamiam skrypt kontaktow...\r\n"
        "python skrypt.py\r\n"
        "pause\r\n"
    )
    try:
        with open(SCIEZKA_BAT, "w", encoding="utf-8") as f:
            f.write(tresc)
        logger.info(f"Wygenerowano skrypt.bat")
    except Exception as e:
        logger.error(f"Blad generowania skrypt.bat: {e}")


# =============================================================================
# KLASA SCRAPERA
# =============================================================================

class SkryptScraper:
    """Scraper danych kontaktowych firm z uzyciem Selenium i BeautifulSoup."""

    def __init__(self, headless: bool = True, timeout: int = 15) -> None:
        self.headless = headless
        self.timeout = timeout
        self.driver: Optional[webdriver.Chrome] = None
        self._lock = threading.Lock()

    def uruchom_przegladarke(self) -> webdriver.Chrome:
        """Tworzy i zwraca nowa instancje przegladarki Chrome."""
        chrome_options = Options()
        if self.headless:
            chrome_options.add_argument("--headless=new")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
        chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
        chrome_options.add_experimental_option(
            "prefs",
            {
                "credentials_enable_service": False,
                "profile.password_manager_enabled": False,
            },
        )

        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_page_load_timeout(self.timeout)
        return driver

    def pobierz_strone(self, driver: webdriver.Chrome, url: str) -> Optional[BeautifulSoup]:
        """Pobiera zawartosc strony i zwraca obiekt BeautifulSoup."""
        if not url or not url.strip():
            return None
        url = url.strip()
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
        for proba in range(RETRIES):
            try:
                driver.get(url)
                time.sleep(1)
                return BeautifulSoup(driver.page_source, "html.parser")
            except TimeoutException:
                logger.debug(f"Timeout dla {url} (proba {proba + 1}/{RETRIES})")
            except WebDriverException as e:
                logger.debug(f"Blad Selenium dla {url}: {str(e)[:80]}")
            except Exception as e:
                logger.debug(f"Blad pobierania {url}: {str(e)[:80]}")
            if proba < RETRIES - 1:
                time.sleep(OPOZNIENIE_MIN)
        return None

    # ----- EKSTRAKCJA DANYCH -----

    @staticmethod
    def wyodrebnij_facebook(soup: BeautifulSoup) -> Optional[str]:
        """Wyodrebnia link do Facebooka."""
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if "facebook.com" in href.lower() and href.startswith("http"):
                return href
        return None

    @staticmethod
    def wyodrebnij_linkedin(soup: BeautifulSoup) -> Optional[str]:
        """Wyodrebnia link do LinkedIn."""
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if "linkedin.com" in href.lower() and href.startswith("http"):
                return href
        return None

    @staticmethod
    def wyodrebnij_email(soup: BeautifulSoup) -> Optional[str]:
        """Wyodrebnia adres email."""
        wzorzec = r"\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b"
        wykluczone = [
            "example.com", "test.com", "domain.com", "yoursite.com",
            "sentry.io", "wix.com",
        ]
        prefiksy_wykluczone = ["noreply", "no-reply", "donotreply"]
        priorytetowe = [
            "kontakt", "biuro", "sekretariat", "info", "office", "contact",
            "handel", "marketing", "recepcja",
        ]

        emaile: List[str] = []

        # mailto:
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if href.startswith("mailto:"):
                email = href.replace("mailto:", "").split("?")[0].strip()
                if re.match(wzorzec, email):
                    emaile.append(email.lower())

        # Wzorzec w tekscie
        tekst = soup.get_text()
        for e in re.findall(wzorzec, tekst):
            emaile.append(e.lower())

        # Filtracja
        unikalne: List[str] = []
        widziane = set()
        for e in emaile:
            if e in widziane:
                continue
            widziane.add(e)
            if any(d in e for d in wykluczone):
                continue
            if any(e.startswith(p) for p in prefiksy_wykluczone):
                continue
            if len(e) > 100:
                continue
            unikalne.append(e)

        # Priorytet
        for e in unikalne:
            if any(kw in e for kw in priorytetowe):
                return e
        return unikalne[0] if unikalne else None

    @staticmethod
    def wyodrebnij_telefon(soup: BeautifulSoup) -> Optional[str]:
        """Wyodrebnia numer telefonu."""
        telefony: List[str] = []

        # Linki tel:
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if href.startswith("tel:"):
                numer = re.sub(r"[^\d+]", "", href.replace("tel:", ""))
                if len(numer) >= 9:
                    telefony.append(numer)

        # Wzorzec w tekscie - polskie numery
        tekst = soup.get_text()
        wzorce = [
            r"(?:\+48[\s\-]?)?\d{3}[\s\-]?\d{3}[\s\-]?\d{3}",
            r"(?:\+48[\s\-]?)?\(\d{2}\)[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}",
            r"\d{2}[\s\-]\d{3}[\s\-]\d{2}[\s\-]\d{2}",
        ]
        for wzorzec in wzorce:
            for m in re.findall(wzorzec, tekst):
                numer = re.sub(r"[^\d+]", "", m)
                if len(numer) >= 9:
                    telefony.append(numer)

        return telefony[0] if telefony else None

    @staticmethod
    def wyodrebnij_nip(soup: BeautifulSoup) -> Optional[str]:
        """Wyodrebnia NIP (10 cyfr)."""
        tekst = soup.get_text()

        # Szukaj wzorca NIP: XXX-XXX-XX-XX lub XXX-XX-XX-XXX lub 10 cyfr
        wzorce = [
            r"NIP[\s:]*(\d{3}[\-\s]?\d{3}[\-\s]?\d{2}[\-\s]?\d{2})",
            r"NIP[\s:]*(\d{3}[\-\s]?\d{2}[\-\s]?\d{2}[\-\s]?\d{3})",
            r"NIP[\s:]*(\d{10})",
        ]
        for wzorzec in wzorce:
            m = re.search(wzorzec, tekst, re.IGNORECASE)
            if m:
                return re.sub(r"[^\d]", "", m.group(1))

        # Szukaj w elementach z klasa/id zawierajacym "nip"
        for elem in soup.find_all(attrs={"class": re.compile(r"nip", re.I)}):
            cyfry = re.sub(r"[^\d]", "", elem.get_text())
            if len(cyfry) == 10:
                return cyfry
        for elem in soup.find_all(attrs={"id": re.compile(r"nip", re.I)}):
            cyfry = re.sub(r"[^\d]", "", elem.get_text())
            if len(cyfry) == 10:
                return cyfry

        return None

    @staticmethod
    def wyodrebnij_osobe_kontaktowa(soup: BeautifulSoup) -> Tuple[Optional[str], Optional[str]]:
        """Wyodrebnia imie, nazwisko i stanowisko osoby kontaktowej.

        Returns:
            Tuple (imie_nazwisko, stanowisko)
        """
        slowa_kontakt = [
            "kontakt", "osoba do kontaktu", "osoba kontaktowa",
            "handel", "marketing", "biuro", "sprzedaz",
            "kierownik", "dyrektor", "prezes", "manager",
            "wlasciciel", "sekretariat",
        ]
        stanowiska = [
            "kierownik", "dyrektor", "prezes", "manager", "specjalista",
            "koordynator", "wlasciciel", "zastepca", "handlowiec",
            "przedstawiciel", "sekretariat", "asystent",
        ]

        wzorzec_imienia = (
            r"\b[A-ZACELNOSZZZ\u0104\u0106\u0118\u0141\u0143\u00d3\u015a\u0179\u017b]"
            r"[a-z\u0105\u0107\u0119\u0142\u0144\u00f3\u015b\u017a\u017c]{2,}"
            r"\s+"
            r"[A-Z\u0104\u0106\u0118\u0141\u0143\u00d3\u015a\u0179\u017b]"
            r"[a-z\u0105\u0107\u0119\u0142\u0144\u00f3\u015b\u017a\u017c]{2,}\b"
        )

        # 1. Szukaj w sekcjach kontaktowych
        for sekcja in soup.find_all(["div", "section", "article", "p", "span"]):
            tekst_sekcji = sekcja.get_text()
            tekst_lower = tekst_sekcji.lower()
            if any(kw in tekst_lower for kw in slowa_kontakt):
                imiona = re.findall(wzorzec_imienia, tekst_sekcji)
                wykluczone = ["Lorem Ipsum", "Jan Kowalski", "Anna Nowak"]
                poprawne = [i for i in imiona if i not in wykluczone]
                if poprawne:
                    stanowisko = None
                    for s in stanowiska:
                        if s in tekst_lower:
                            stanowisko = s.capitalize()
                            break
                    return poprawne[0], stanowisko

        # 2. Szukaj w naglowkach sekcji kontaktowych
        for naglowek in soup.find_all(["h1", "h2", "h3", "h4"]):
            tekst_nag = naglowek.get_text().lower()
            if any(kw in tekst_nag for kw in slowa_kontakt):
                nastepny = naglowek.find_next(["p", "div", "span"])
                if nastepny:
                    imiona = re.findall(wzorzec_imienia, nastepny.get_text())
                    if imiona:
                        return imiona[0], None

        # 3. Wzorzec "Osoba: Imie Nazwisko"
        wzorzec_kontakt = (
            r"(?:kontakt|osoba|kierownik|dyrektor|prezes)"
            r"[\s:]+"
            r"([A-Z\u0104\u0106\u0118\u0141\u0143\u00d3\u015a\u0179\u017b]"
            r"[a-z\u0105\u0107\u0119\u0142\u0144\u00f3\u015b\u017a\u017c]+"
            r"\s+"
            r"[A-Z\u0104\u0106\u0118\u0141\u0143\u00d3\u015a\u0179\u017b]"
            r"[a-z\u0105\u0107\u0119\u0142\u0144\u00f3\u015b\u017a\u017c]+)"
        )
        m = re.search(wzorzec_kontakt, soup.get_text(), re.IGNORECASE)
        if m:
            return m.group(1), None

        return None, None

    def scrapuj_firme(
        self, driver: webdriver.Chrome, nazwa: str, www: str
    ) -> Dict[str, Optional[str]]:
        """Scrapuje dane kontaktowe dla jednej firmy."""
        wynik: Dict[str, Optional[str]] = {
            "facebook": None,
            "linkedin": None,
            "email": None,
            "telefon": None,
            "nip": None,
            "imie_nazwisko": None,
            "stanowisko": None,
        }

        if not www or not www.strip():
            return wynik

        www = www.strip()
        if not www.startswith(("http://", "https://")):
            www = "https://" + www

        # Lista podstron do sprawdzenia
        podstrony = [
            www,
            urljoin(www, "/kontakt"),
            urljoin(www, "/kontakt.html"),
            urljoin(www, "/contact"),
            urljoin(www, "/contact.html"),
            urljoin(www, "/o-nas"),
            urljoin(www, "/o-firmie"),
            urljoin(www, "/about"),
            urljoin(www, "/zespol"),
            urljoin(www, "/team"),
        ]

        for url in podstrony:
            soup = self.pobierz_strone(driver, url)
            if not soup:
                continue

            if not wynik["facebook"]:
                wynik["facebook"] = self.wyodrebnij_facebook(soup)
            if not wynik["linkedin"]:
                wynik["linkedin"] = self.wyodrebnij_linkedin(soup)
            if not wynik["email"]:
                wynik["email"] = self.wyodrebnij_email(soup)
            if not wynik["telefon"]:
                wynik["telefon"] = self.wyodrebnij_telefon(soup)
            if not wynik["nip"]:
                wynik["nip"] = self.wyodrebnij_nip(soup)
            if not wynik["imie_nazwisko"]:
                osoba, stanowisko = self.wyodrebnij_osobe_kontaktowa(soup)
                if osoba:
                    wynik["imie_nazwisko"] = osoba
                    wynik["stanowisko"] = stanowisko

            # Jesli znaleziono kluczowe dane, przerwij
            if wynik["email"] and wynik["imie_nazwisko"]:
                break

            time.sleep(
                OPOZNIENIE_MIN + (OPOZNIENIE_MAX - OPOZNIENIE_MIN) * (hash(url) % 100 / 100.0)
            )

        return wynik

    def szukaj_duckduckgo(
        self, driver: webdriver.Chrome, nazwa: str, brakujace: List[str]
    ) -> Dict[str, Optional[str]]:
        """Szuka brakujacych danych firmy w DuckDuckGo."""
        wynik: Dict[str, Optional[str]] = {}
        zapytania: List[str] = []

        if "facebook" in brakujace:
            zapytania.append(f'"{nazwa}" facebook')
        if "linkedin" in brakujace:
            zapytania.append(f'"{nazwa}" linkedin')
        if "email" in brakujace:
            zapytania.append(f'"{nazwa}" kontakt email')
        if "telefon" in brakujace:
            zapytania.append(f'"{nazwa}" telefon kontakt')
        if "nip" in brakujace:
            zapytania.append(f'"{nazwa}" NIP')

        for zapytanie in zapytania[:MAX_WYSZUKIWAN_DDG]:
            try:
                driver.get("https://duckduckgo.com/")
                time.sleep(1.5)

                pole = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.NAME, "q"))
                )
                pole.clear()
                pole.send_keys(zapytanie)
                time.sleep(0.5)
                pole.send_keys(Keys.RETURN)
                time.sleep(2)

                soup = BeautifulSoup(driver.page_source, "html.parser")

                if "facebook" in brakujace and "facebook" not in wynik:
                    fb = self.wyodrebnij_facebook(soup)
                    if fb:
                        wynik["facebook"] = fb
                        brakujace.remove("facebook")

                if "linkedin" in brakujace and "linkedin" not in wynik:
                    li = self.wyodrebnij_linkedin(soup)
                    if li:
                        wynik["linkedin"] = li
                        brakujace.remove("linkedin")

                if "email" in brakujace and "email" not in wynik:
                    em = self.wyodrebnij_email(soup)
                    if em:
                        wynik["email"] = em
                        brakujace.remove("email")

                time.sleep(OPOZNIENIE_MIN)

            except Exception as e:
                logger.debug(f"Blad DuckDuckGo dla '{zapytanie}': {str(e)[:80]}")
                continue

        return wynik


# =============================================================================
# PRZETWARZANIE JEDNEJ FIRMY (WATEK)
# =============================================================================

def przetworz_firme(
    wiersz: int,
    nazwa: str,
    www: str,
    scraper: SkryptScraper,
) -> Tuple[int, Dict[str, Optional[str]]]:
    """Przetwarza jedna firme - uruchamiana w watku."""
    driver = None
    try:
        driver = scraper.uruchom_przegladarke()
        wynik = scraper.scrapuj_firme(driver, nazwa, www)

        # Sprawdz brakujace dane i szukaj w DuckDuckGo
        brakujace = [k for k in ["facebook", "linkedin", "email"] if not wynik.get(k)]
        if brakujace and nazwa:
            ddg = scraper.szukaj_duckduckgo(driver, nazwa, brakujace)
            for klucz, wartosc in ddg.items():
                if not wynik.get(klucz):
                    wynik[klucz] = wartosc

        return wiersz, wynik
    except Exception as e:
        logger.error(f"Blad przetwarzania wiersza {wiersz} ({nazwa}): {e}")
        return wiersz, {
            "facebook": None,
            "linkedin": None,
            "email": None,
            "telefon": None,
            "nip": None,
            "imie_nazwisko": None,
            "stanowisko": None,
        }
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


# =============================================================================
# GLOWNA FUNKCJA PRZETWARZANIA
# =============================================================================

def przetwarzaj_baze() -> None:
    """Glowna funkcja przetwarzajaca baze Excel."""

    # --- Sprawdzenie pliku ---
    if not SCIEZKA_EXCEL.exists():
        mow(f"BLAD: Nie znaleziono pliku: {SCIEZKA_EXCEL}")
        mow(f"Umiesc plik {NAZWA_ARKUSZA} w folderze: {FOLDER_SKRYPTU}")
        return

    # --- Wczytanie Excela ---
    mow(f"Wczytuje plik: {NAZWA_ARKUSZA}")
    try:
        wb = openpyxl.load_workbook(str(SCIEZKA_EXCEL))
        ws = wb.active
    except Exception as e:
        mow(f"BLAD: Nie mozna otworzyc pliku Excel: {e}")
        return

    total = ws.max_row - 1  # Pomijamy naglowek
    if total <= 0:
        mow("BLAD: Plik Excel jest pusty lub ma tylko naglowek.")
        return

    mow(f"Znaleziono {total} firm w pliku.")

    # --- Checkpoint ---
    wiersz_start = WIERSZ_STARTOWY
    ck = wczytaj_checkpoint()
    if ck is not None and ck > wiersz_start:
        odp = zapytaj(
            f"Znaleziono checkpoint: wiersz {ck}. Wznowic od wiersza {ck + 1}? (T/N, ENTER=tak)"
        )
        if odp.lower() != "n":
            wiersz_start = ck + 1
            mow(f"Wznawiam od wiersza {wiersz_start}.")
        else:
            mow("Zaczynam od poczatku.")

    # --- Folder wynikow ---
    SCIEZKA_WYNIKOW.mkdir(parents=True, exist_ok=True)

    # --- Przygotowanie listy firm ---
    firmy: List[Tuple[int, str, str]] = []
    for wiersz in range(wiersz_start, ws.max_row + 1):
        nazwa = ws.cell(row=wiersz, column=KOL_NAZWA_FIRMY).value
        www = ws.cell(row=wiersz, column=KOL_WWW).value
        if nazwa:
            firmy.append((wiersz, str(nazwa), str(www) if www else ""))

    if not firmy:
        mow("Brak firm do przetworzenia.")
        return

    mow(f"Do przetworzenia: {len(firmy)} firm, {MAX_WATKOW} watkow.")

    odp = zapytaj("Nacisnij ENTER aby rozpoczac, Q aby wyjsc.")
    if odp.lower() == "q":
        mow("Przerwano przez uzytkownika.")
        return

    # --- Scraper ---
    scraper = SkryptScraper(headless=HEADLESS, timeout=TIMEOUT_STRONA)
    przetworzone = 0
    bledy = 0

    zapisz_log_csv("Start przetwarzania")

    # --- Przetwarzanie porcjami ---
    for i in range(0, len(firmy), CHECKPOINT_CO):
        porcja = firmy[i : i + CHECKPOINT_CO]

        with ThreadPoolExecutor(max_workers=MAX_WATKOW) as executor:
            futures = {
                executor.submit(przetworz_firme, w, n, u, scraper): (w, n)
                for w, n, u in porcja
            }

            for future in as_completed(futures):
                wiersz_nr, nazwa_f = futures[future]
                try:
                    nr, dane = future.result()

                    # Zapisz wyniki do Excela
                    if dane.get("facebook"):
                        ws.cell(row=nr, column=KOL_FACEBOOK, value=dane["facebook"])
                    if dane.get("linkedin"):
                        ws.cell(row=nr, column=KOL_LINKEDIN, value=dane["linkedin"])
                    if dane.get("email"):
                        ws.cell(row=nr, column=KOL_EMAIL, value=dane["email"])
                    if dane.get("telefon"):
                        ws.cell(row=nr, column=KOL_TELEFON, value=dane["telefon"])
                    if dane.get("nip"):
                        ws.cell(row=nr, column=KOL_NIP, value=dane["nip"])
                    if dane.get("imie_nazwisko"):
                        ws.cell(row=nr, column=KOL_IMIE_NAZWISKO, value=dane["imie_nazwisko"])
                    if dane.get("stanowisko"):
                        ws.cell(row=nr, column=KOL_STANOWISKO, value=dane["stanowisko"])

                    przetworzone += 1
                    znalezione = sum(1 for v in dane.values() if v)
                    zapisz_log_csv(f"Przetworzono {nazwa_f} | znaleziono {znalezione} pol")

                except Exception as e:
                    bledy += 1
                    logger.error(f"Blad wiersza {wiersz_nr}: {e}")
                    zapisz_log_csv(f"BLAD {nazwa_f} | {str(e)[:60]}")

                mow_postep(przetworzone + bledy, len(firmy), nazwa_f)

        # Checkpoint
        ostatni = porcja[-1][0]
        zapisz_checkpoint(ostatni)

        # Zapis posredni
        try:
            plik_wynikowy = SCIEZKA_WYNIKOW / f"wyniki_{NAZWA_ARKUSZA}"
            wb.save(str(plik_wynikowy))
        except Exception as e:
            logger.error(f"Blad zapisu posredniego: {e}")

    # --- Zapis koncowy ---
    plik_wynikowy = SCIEZKA_WYNIKOW / f"wyniki_{NAZWA_ARKUSZA}"
    try:
        wb.save(str(plik_wynikowy))
        mow(f"Wyniki zapisane do: {plik_wynikowy}")
    except Exception as e:
        mow(f"BLAD zapisu wynikow: {e}")

    # Usun checkpoint po zakonczeniu
    try:
        if SCIEZKA_CHECKPOINT.exists():
            SCIEZKA_CHECKPOINT.unlink()
    except Exception:
        pass

    # --- Podsumowanie ---
    zapisz_log_csv(f"Koniec | przetworzone={przetworzone} bledy={bledy}")

    mow("=" * 40)
    mow("PODSUMOWANIE")
    mow(f"Przetworzone firmy: {przetworzone}")
    mow(f"Bledy: {bledy}")
    mow(f"Plik wynikowy: {plik_wynikowy}")
    mow("=" * 40)


# =============================================================================
# PUNKT WEJSCIA
# =============================================================================

if __name__ == "__main__":
    mow("SKRYPT UZUPELNIANIA BAZY KONTAKTOW FIRM")
    mow("Wersja 3.0 - NVDA-friendly")

    # Generuj pliki konfiguracyjne
    generuj_config_json()
    generuj_bat()

    mow(f"Plik wejsciowy: {NAZWA_ARKUSZA}")
    mow(f"Folder wynikow: {SCIEZKA_WYNIKI}")
    mow(f"Watki: {MAX_WATKOW}")
    mow(f"Tryb headless: {'tak' if HEADLESS else 'nie'}")

    przetwarzaj_baze()

    mow("Skrypt zakonczony.")
    zapytaj("Nacisnij ENTER aby zamknac.")
