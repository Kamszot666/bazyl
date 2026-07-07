"""
=============================================================================
SCRAPER BAZY DANYCH — SZKIELET SKRYPTU
=============================================================================
Cel:    Automatyczne pobieranie danych organizacji/firm z Internetu
        i zapis do pliku XLSX zgodnego z formatką firmy.

Użycie: python scraper.py
        python scraper.py --reset    ← usuwa postęp i zaczyna od nowa

Wznowienie po awarii: uruchom ponownie — skrypt czyta progress.json
                      i kontynuuje od miejsca zatrzymania.

Rekonfiguracja pod nowe zadanie: zmień wyłącznie blok CONFIG poniżej.
=============================================================================
"""

import argparse
import json
import logging
import random
import re
import sys
import time
import urllib.parse
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import httpx
import phonenumbers
import requests
import pdfplumber
from bs4 import BeautifulSoup
from io import BytesIO
from openpyxl import Workbook, load_workbook
from selectolax.parser import HTMLParser
from openpyxl.styles import Alignment, Font, PatternFill
from tqdm import tqdm


# =============================================================================
# KONFIGURACJA — zmień tylko tę sekcję przy nowym zadaniu
# =============================================================================
CONFIG = {
    # --- IDENTYFIKACJA ZADANIA ---
    "task_name": "organizacje_polonijne",
    "output_file": "output/baza_organizacje_polonijne.xlsx",
    "rejected_file": "output/odrzucone.xlsx",
    "log_file": "scraper_log.txt",
    "progress_file": "progress.json",

    # --- ETAP 1: ZBIERANIE NAZW OFERENTÓW Z WYNIKÓW KONKURSÓW MSZ/KPRM ---
    # Plik pośredni — tylko nazwa organizacji, projekt, kwota i status.
    # Adres korespondencyjny, KRS, telefon, e-mail itd. uzupełnia dopiero
    # kolejny etap (wzbogacanie przez API KRS/REGON), którego jeszcze nie ma.
    "seed_file": "output/oferenci_konkursy_polonijne.xlsx",

    # --- ETAP 2: MAPA NAZWA OFERENTA -> NUMER KRS (weryfikacja ręczna) ---
    # Plik zawiera wynik ręcznego sprawdzenia każdej z 404 unikalnych nazw
    # oferentów przez oficjalne API KRS (patrz historia rozmowy). Każdy
    # numer KRS występuje w pliku dokładnie raz (duplikaty scalone).
    "mapa_krs_file": "output/mapa_krs_oferentow.json",

    # --- ETAP 3: WZBOGACANIE DANYCH PRZEZ OFICJALNE API KRS ---
    # Cache odpowiedzi API KRS wg numeru KRS — pozwala wznowić Etap 3 po
    # awarii bez ponownego odpytywania już pobranych numerów.
    "cache_dane_krs_file": "output/cache_dane_krs.json",
    "api_krs_url_wzor": "https://api-krs.ms.gov.pl/api/krs/OdpisAktualny/{krs}?rejestr={rejestr}&format=json",

    # --- ETAP 4: DANE KONTAKTOWE ZE STRON WWW ORGANIZACJI ---
    # Adres strony każdej organizacji ustala biblioteka ddgs (pierwsza próba)
    # albo wyszukiwanie ręczne (zapasowa metoda) — patrz CLAUDE.md. Jeśli
    # ddgs zgłosi blokadę/limit zapytań, automatyczne wyszukiwanie wyłącza
    # się do końca uruchomienia i wymaga wyszukania ręcznego (nigdy nie
    # próbujemy obchodzić zabezpieczeń antybotowych).
    # Pobranie znanego adresu i wyciągnięcie z niego telefonu, e-maila i
    # profilu social media odbywa się uniwersalnymi wzorcami tekstowymi
    # (nie selektorami CSS specyficznymi dla konkretnej strony).
    "dane_kontaktowe_file": "output/dane_kontaktowe.json",
    "podstrony_kontaktowe": ["kontakt", "kontakty", "o-nas", "contact"],
    "domeny_social_media": [
        "facebook.com", "instagram.com", "linkedin.com",
        "youtube.com", "x.com", "twitter.com",
    ],
    # Próg długości tekstu strony (po pobraniu httpx), poniżej którego
    # uznajemy stronę za renderowaną przez JavaScript i sięgamy po
    # Playwright jako zapasowe narzędzie.
    "prog_dlugosci_tekstu_js": 200,

    # Każdy adres poniżej zweryfikowano ręcznie przez realne pobranie strony
    # (patrz historia rozmowy — nazwa i podział konkursu zmienia się rok do
    # roku, więc NIE da się zbudować jednego wzorca adresu zależnego od roku).
    "konkursy_polonijne": [
        {
            "url": "https://www.gov.pl/web/dyplomacja/wyniki-konkursu-polonia-i-polacy-za-granica-2026",
            "rok": 2026,
            "nazwa_konkursu": "Polonia i Polacy za Granicą 2026",
            "archiwum_wayback": False,
        },
        {
            "url": "https://www.gov.pl/web/dyplomacja/ogloszenie-wynikow-konkursu-wsparcie-srodowisk-polonijnych-w-obszarze-aktywizacji-postaw-obywatelskich",
            "rok": 2025,
            "nazwa_konkursu": "Wsparcie środowisk polonijnych w obszarze aktywizacji postaw obywatelskich 2025",
            "archiwum_wayback": False,
        },
        {
            "url": "https://www.gov.pl/web/dyplomacja/wyniki-konkursu-infrastruktura-polonijna-2025",
            "rok": 2025,
            "nazwa_konkursu": "Infrastruktura Polonijna 2025",
            "archiwum_wayback": False,
        },
        {
            "url": "https://www.gov.pl/web/dyplomacja/wyniki-konkursu-polonia-i-polacy-za-granica-2024--media-i-struktury2",
            "rok": 2024,
            "nazwa_konkursu": "Polonia i Polacy za Granicą 2024 – Media i Struktury",
            "archiwum_wayback": False,
        },
        {
            "url": "https://www.gov.pl/web/dyplomacja/ogloszenie-wynikow-konkursu-polonia-i-polacy-za-granica-2024---wydarzenia-i-inicjatywy-polonijne",
            "rok": 2024,
            "nazwa_konkursu": "Polonia i Polacy za Granicą 2024 – Wydarzenia i inicjatywy polonijne",
            "archiwum_wayback": False,
        },
        {
            "url": "https://www.gov.pl/web/dyplomacja/ogloszenie-wynikow-konkursu-polonia-i-polacy-za-granica-2024---regranting",
            "rok": 2024,
            "nazwa_konkursu": "Polonia i Polacy za Granicą 2024 – Regranting",
            "archiwum_wayback": False,
        },
        {
            # Strona oryginalna KPRM już nie istnieje (kompetencje przejęło MSZ
            # w połowie 2024 roku) — dane pobieramy z Wayback Machine.
            "url": "https://web.archive.org/web/20240201084241/https://www.gov.pl/web/polonia/wyniki-konkursu-polonia-i-polacy-za-granica-2023",
            "rok": 2023,
            "nazwa_konkursu": "Polonia i Polacy za Granicą 2023",
            "archiwum_wayback": True,
        },
    ],

    # Załączniki o tytułach zawierających te frazy pomijamy — to zbiorcze
    # listy "wszystkich zgłoszonych ofert", które nie mówią nic o wyniku
    # (dublują się z listą przyznanych + listą odrzuconych/nieprzyjętych).
    "zalaczniki_pomijane_frazy": ["zgłoszon", "wpłynęł"],

    # --- ŹRÓDŁA DANYCH (w kolejności priorytetu) ---
    # Uzupełnij przed uruchomieniem po konsultacji z użytkownikiem.
    # Każde źródło: url (punkt startowy), priority (1=najwyższy), type, notes
    "sources": [
        {
            "id": "ngo_pl",
            "url": "https://baza.ngo.pl/szukaj/organizacje?query=polonijna",
            "priority": 1,
            "type": "authoritative",
            "requires_api_key": False,
            "requires_login": False,
            "rate_limit": "1 req/s",
            "notes": "Główna autorytatywna baza NGO w Polsce",
        },
        # Dodaj kolejne źródła według potrzeb:
        # {
        #     "id": "gov_portal",
        #     "url": "https://...",
        #     "priority": 2,
        #     "type": "supplementary",
        #     "requires_api_key": False,
        #     "requires_login": False,
        #     "rate_limit": "2 req/s",
        #     "notes": "Portal rządowy — uzupełniający",
        # },
    ],

    # --- FILTROWANIE ---
    "filter_country": "Polska",           # siedziba musi być w Polsce
    "filter_active_months": 12,           # aktywna w ciągu ostatnich N miesięcy
    "filter_exclude_sole_trader": False,  # True = wyklucz jednoosobowe działalności

    # --- LIMITY DANYCH ---
    "max_phones_org": 3,                  # maks. telefonów do organizacji
    "max_emails_org": 3,                  # maks. e-maili do organizacji

    # --- SIEĆ ---
    "delay_min": 1.0,                     # minimalne opóźnienie między req (sekundy)
    "delay_max": 3.0,                     # maksymalne opóźnienie między req (sekundy)
    "retry_delay": 5.0,                   # czas przed ponowną próbą po błędzie
    "request_timeout": 15,                # timeout jednego żądania HTTP (sekundy)

    # --- WYJŚCIE ---
    "encoding": "utf-8-sig",             # UTF-8 z BOM = poprawne polskie znaki w Excel
    "date_format": "%d.%m.%Y",           # format daty w kolumnie date_collected
}

# Nagłówek HTTP — symulujemy normalną przeglądarkę
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pl-PL,pl;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Kolumny pliku wyjściowego (bez Lp. — użytkownik wypełnia ręcznie)
OUTPUT_COLUMNS = [
    "Kategoria",
    "Branża / Typ",
    "Nazwa",
    "Adres korespondencyjny",
    "Województwo",
    "Numer telefonu",
    "Adres e-mail",
    "Strona WWW",
    "Profil w mediach społecznościowych",
    "Osoba kontaktowa",
    "Numer telefonu do osoby kontaktowej",
    "Adres e-mail do osoby kontaktowej",
    "Data pozyskania informacji",
    "Krótka charakterystyka podmiotu",
    # Kolumny dodatkowe (poza formatką firmy):
    "URL źródła",
    "KRS",
    "REGON",
    "NIP",
]

# Kolumny pliku pośredniego z etapu 1 (zbieranie oferentów z konkursów).
# To NIE jest jeszcze formatka — brakuje adresu, KRS, kontaktu itd.
SEED_COLUMNS = [
    "Rok",
    "Nazwa konkursu",
    "Lista źródłowa (status)",
    "Oferent",
    "Nazwa projektu",
    "Kwota",
    "URL artykułu",
    "URL załącznika PDF",
]


# =============================================================================
# LOGOWANIE
# =============================================================================

def setup_logging():
    """Konfiguruje logging do pliku i konsoli."""
    logger = logging.getLogger("scraper")
    logger.setLevel(logging.DEBUG)

    # Formatka: zdarzenie | status | czas
    class PipeFormatter(logging.Formatter):
        def format(self, record):
            status = "ERROR" if record.levelno >= logging.ERROR else (
                "INFO" if record.levelno == logging.INFO else "SUCCESS"
            )
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return f"{record.getMessage()} | {status} | {ts}"

    # Plik logów (dopisywanie, nie nadpisywanie)
    fh = logging.FileHandler(CONFIG["log_file"], mode="a", encoding="utf-8")
    fh.setFormatter(PipeFormatter())
    logger.addHandler(fh)

    # Konsola (czytelniejszy format)
    ch = logging.StreamHandler()
    ch.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
    logger.addHandler(ch)

    return logger


log = setup_logging()


def log_event(event_name: str, success: bool, details: str = ""):
    """Zapisuje zdarzenie do logu w formacie: zdarzenie | SUCCESS/ERROR | czas"""
    msg = event_name if not details else f"{event_name} — {details}"
    if success:
        log.info(msg)
    else:
        log.error(msg)


# =============================================================================
# POSTĘP I WZNOWIENIE
# =============================================================================

def load_progress() -> dict:
    """Wczytuje stan postępu z pliku JSON."""
    p = Path(CONFIG["progress_file"])
    if p.exists():
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {"completed_ids": [], "last_source": None, "total_scraped": 0}


def save_progress(progress: dict):
    """Zapisuje aktualny stan postępu."""
    with open(CONFIG["progress_file"], "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def reset_progress():
    """Usuwa plik postępu — wymusza start od początku."""
    p = Path(CONFIG["progress_file"])
    if p.exists():
        p.unlink()
        log_event("RESET_POSTEPU", True)


# =============================================================================
# SIEĆ — pobieranie stron
# =============================================================================

def fetch_page(url: str, session: requests.Session) -> BeautifulSoup | None:
    """
    Pobiera stronę i zwraca BeautifulSoup lub None po nieudanych próbach.
    Dwie próby: pierwsza → czeka retry_delay → druga → None.
    """
    for attempt in range(1, 3):
        try:
            resp = session.get(url, headers=HEADERS, timeout=CONFIG["request_timeout"])
            resp.raise_for_status()
            resp.encoding = "utf-8"
            log_event(f"POBIERANIE {url[:60]}", True)
            return BeautifulSoup(resp.text, "lxml")

        except requests.RequestException as e:
            log_event(f"POBIERANIE {url[:60]} (próba {attempt})", False, str(e))
            if attempt < 2:
                time.sleep(CONFIG["retry_delay"])

    return None  # obie próby nieudane


def polite_delay():
    """Losowe opóźnienie między żądaniami — szanujemy serwer."""
    time.sleep(random.uniform(CONFIG["delay_min"], CONFIG["delay_max"]))


# =============================================================================
# PARSOWANIE — dostosuj do każdego źródła
# =============================================================================

def get_entity_urls_from_source(source: dict, session: requests.Session) -> list[str]:
    """
    Pobiera listę URL-i do poszczególnych rekordów ze źródła.

    !! DOSTOSUJ TĘ FUNKCJĘ DO KAŻDEGO ŹRÓDŁA !!

    Przykład dla baza.ngo.pl:
      - pobierz stronę listową
      - znajdź linki do profili organizacji
      - obsłuż paginację (kolejne strony)
    """
    urls = []
    page_url = source["url"]

    while page_url:
        soup = fetch_page(page_url, session)
        if soup is None:
            log_event(f"LISTA_{source['id'].upper()}", False, f"Nie można pobrać: {page_url}")
            break

        # ---- DOSTOSUJ: selektory CSS/XPath dla konkretnego źródła ----
        # Przykład (zastąp prawdziwymi selektorami):
        # links = soup.select("a.organization-link")
        # for link in links:
        #     href = link.get("href", "")
        #     if href:
        #         urls.append(href if href.startswith("http") else BASE_URL + href)

        # Przykład paginacji:
        # next_btn = soup.select_one("a.next-page")
        # page_url = next_btn["href"] if next_btn else None

        # Tymczasowy placeholder — usuń po implementacji:
        log.warning(f"get_entity_urls_from_source: zaimplementuj parsowanie dla źródła '{source['id']}'")
        page_url = None  # usuń gdy paginacja będzie gotowa
        polite_delay()

    log_event(f"LISTA_{source['id'].upper()}", True, f"Znaleziono {len(urls)} URL-i")
    return urls


def parse_entity_page(url: str, soup: BeautifulSoup) -> dict:
    """
    Parsuje stronę jednej organizacji/firmy i zwraca słownik danych.

    !! DOSTOSUJ TĘ FUNKCJĘ DO STRUKTURY STRONY ŹRÓDŁOWEJ !!

    Zwraca dict z kluczami odpowiadającymi OUTPUT_COLUMNS.
    Jeśli danych brak → wartość "none" (lub "brak" / "nie ustalono" zgodnie ze spec.).
    """
    today = datetime.now().strftime(CONFIG["date_format"])

    # ---- DOSTOSUJ: selektory dla konkretnej strony ----
    # Poniżej schemat — zastąp rzeczywistymi selektorami:

    record = {
        "Kategoria":             extract_text(soup, "span.category") or "organizacja pozarządowa",
        "Branża / Typ":          extract_text(soup, "span.industry") or "none",
        "Nazwa":                 extract_text(soup, "h1.org-name") or "none",
        "Adres korespondencyjny": extract_address(soup),
        "Województwo":           extract_text(soup, "span.voivodeship") or "none",
        "Numer telefonu":        extract_phones(soup, max_count=CONFIG["max_phones_org"]),
        "Adres e-mail":          extract_emails(soup, max_count=CONFIG["max_emails_org"]),
        "Strona WWW":            extract_website(soup),
        "Profil w mediach społecznościowych": extract_social(soup),
        "Osoba kontaktowa":      extract_text(soup, "span.contact-person") or "nie ustalono",
        "Numer telefonu do osoby kontaktowej": extract_text(soup, "span.contact-phone") or "brak",
        "Adres e-mail do osoby kontaktowej":   extract_text(soup, "span.contact-email") or "brak",
        "Data pozyskania informacji": today,
        "Krótka charakterystyka podmiotu": extract_description(soup),
        # Kolumny dodatkowe:
        "URL źródła": url,
        "KRS":   extract_registry_number(soup, "krs") or "none",
        "REGON": extract_registry_number(soup, "regon") or "none",
        "NIP":   extract_registry_number(soup, "nip") or "none",
    }

    return record


# =============================================================================
# HELPERY PARSOWANIA — dostosuj selektory CSS
# =============================================================================

def extract_text(soup: BeautifulSoup, selector: str) -> str | None:
    """Zwraca tekst pierwszego elementu pasującego do selektora CSS."""
    el = soup.select_one(selector)
    if el:
        return el.get_text(strip=True) or None
    return None


def extract_address(soup: BeautifulSoup) -> str:
    """
    Wyciąga adres korespondencyjny.
    !! DOSTOSUJ SELEKTOR !!
    """
    # Przykład:
    # parts = [
    #     extract_text(soup, "span.street"),
    #     extract_text(soup, "span.postcode"),
    #     extract_text(soup, "span.city"),
    # ]
    # return ", ".join(p for p in parts if p) or "none"
    return "none"  # zastąp implementacją


def extract_phones(soup: BeautifulSoup, max_count: int = 3) -> str:
    """
    Wyciąga telefony do organizacji (maks. max_count).
    Priorytet: sekretariat > biuro > ogólny.
    Format wynikowy: XXX XXX XXX, kolejne oddzielone \\n
    !! DOSTOSUJ SELEKTOR !!
    """
    # Przykład:
    # raw_phones = [el.get_text(strip=True) for el in soup.select("span.phone")]
    # cleaned = [normalize_phone(p) for p in raw_phones if p]
    # return "\n".join(cleaned[:max_count]) or "none"
    return "none"  # zastąp implementacją


def extract_emails(soup: BeautifulSoup, max_count: int = 3) -> str:
    """
    Wyciąga e-maile do organizacji (maks. max_count).
    Priorytet: sekretariat > biuro > ogólny.
    !! DOSTOSUJ SELEKTOR !!
    """
    # Przykład:
    # raw = [el.get_text(strip=True) for el in soup.select("a[href^='mailto:']")]
    # clean = [e.replace("mailto:", "").strip() for e in raw if "@" in e]
    # # Priorytet: sekretariat > biuro > info > reszta
    # priority = ["sekretariat", "biuro", "info", "kontakt", "office"]
    # sorted_emails = sorted(clean, key=lambda e: next(
    #     (i for i, p in enumerate(priority) if p in e.lower()), len(priority)
    # ))
    # return "\n".join(sorted_emails[:max_count]) or "none"
    return "none"  # zastąp implementacją


def extract_website(soup: BeautifulSoup) -> str:
    """
    Wyciąga URL strony WWW.
    Preferuje podstronę /kontakt lub /contact.
    !! DOSTOSUJ SELEKTOR !!
    """
    # Przykład:
    # link = soup.select_one("a.website-link")
    # if link:
    #     base = link.get("href", "").rstrip("/")
    #     # Sprawdź czy istnieje podstrona kontaktowa
    #     for suffix in ["/kontakt", "/contact", "/kontakty"]:
    #         contact_url = base + suffix
    #         # Opcjonalnie: zweryfikuj istnienie ping-iem (ostrożnie z limitami)
    #         return contact_url
    #     return base
    return "brak"  # zastąp implementacją


def extract_social(soup: BeautifulSoup) -> str:
    """
    Wyciąga URL do profilu w mediach społecznościowych.
    Priorytet: Facebook > LinkedIn > Instagram > inne.
    !! DOSTOSUJ SELEKTOR !!
    """
    # Przykład:
    # priority_domains = ["facebook.com", "linkedin.com", "instagram.com", "twitter.com"]
    # all_links = [a.get("href", "") for a in soup.select("a[href]")]
    # for domain in priority_domains:
    #     for link in all_links:
    #         if domain in link:
    #             return link
    return "brak"  # zastąp implementacją


def extract_description(soup: BeautifulSoup) -> str:
    """
    Wyciąga krótką charakterystykę (1–3 zdania).
    !! DOSTOSUJ SELEKTOR !!
    """
    # Przykład:
    # desc = extract_text(soup, "div.org-description")
    # if desc:
    #     sentences = desc.split(". ")
    #     return ". ".join(sentences[:3]).strip() + ("." if not desc.endswith(".") else "")
    return "none"  # zastąp implementacją


def extract_registry_number(soup: BeautifulSoup, reg_type: str) -> str | None:
    """
    Wyciąga numer z rejestru: krs / regon / nip.
    !! DOSTOSUJ SELEKTOR !!
    """
    # Przykład — szukamy tekstu zawierającego etykietę:
    # label_map = {"krs": "KRS", "regon": "REGON", "nip": "NIP"}
    # label = label_map.get(reg_type, "")
    # for row in soup.select("tr, div.data-row"):
    #     text = row.get_text()
    #     if label in text:
    #         numbers = re.findall(r"\d{6,10}", text)
    #         if numbers:
    #             return numbers[0]
    return None  # zastąp implementacją


# Dwucyfrowe numery kierunkowe polskich telefonów stacjonarnych (numeracja
# krajowa UKE) — jeśli pierwsze dwie cyfry numeru pasują do tej listy,
# traktujemy numer jako stacjonarny i zapisujemy z kierunkowym oddzielnie,
# zgodnie z przykładem z procedury ("33 874 22 33"). W przeciwnym razie
# traktujemy numer jako komórkowy (grupowanie XXX XXX XXX).
KIERUNKOWE_STACJONARNE = {
    "12", "13", "14", "15", "16", "17", "18", "22", "23", "24", "25", "29",
    "32", "33", "34", "41", "42", "43", "44", "46", "47", "48", "52", "54",
    "55", "56", "58", "59", "61", "62", "63", "64", "65", "67", "68", "71",
    "74", "75", "76", "77", "81", "82", "83", "84", "85", "86", "87", "89",
    "91", "94", "95",
}


def normalize_phone(raw: str) -> str:
    """Normalizuje numer telefonu do formatu XXX XXX XXX (komórkowy) lub XX XXX XX XX (stacjonarny z kierunkowym)."""
    digits = re.sub(r"\D", "", raw)
    # Usuń polskie prefiksy (+48, 0048)
    if digits.startswith("0048"):
        digits = digits[4:]
    elif digits.startswith("48") and len(digits) == 11:
        digits = digits[2:]

    if len(digits) == 9:
        if digits[0:2] in KIERUNKOWE_STACJONARNE:
            return f"{digits[0:2]} {digits[2:5]} {digits[5:7]} {digits[7:9]}"
        return f"{digits[0:3]} {digits[3:6]} {digits[6:9]}"
    # Inne długości — zwróć cyfry ze spacjami co 3
    if len(digits) >= 7:
        parts = [digits[i:i+3] for i in range(0, len(digits), 3)]
        return " ".join(parts)
    return raw  # zwróć oryginał jeśli nie da się znormalizować


def is_active(last_activity_date: datetime | None) -> bool:
    """Sprawdza czy organizacja była aktywna w ciągu ostatnich N miesięcy."""
    if last_activity_date is None:
        return True  # nie możemy zweryfikować → przyjmujemy że aktywna
    cutoff = datetime.now() - timedelta(days=30 * CONFIG["filter_active_months"])
    return last_activity_date >= cutoff


def is_based_in_poland(record: dict) -> bool:
    """Sprawdza czy siedziba organizacji jest w Polsce."""
    address = (record.get("Adres korespondencyjny") or "").lower()
    voivodeship = (record.get("Województwo") or "").lower()
    # Polskie województwa
    polish_voivodeships = {
        "dolnośląskie", "kujawsko-pomorskie", "lubelskie", "lubuskie",
        "łódzkie", "małopolskie", "mazowieckie", "opolskie", "podkarpackie",
        "podlaskie", "pomorskie", "śląskie", "świętokrzyskie",
        "warmińsko-mazurskie", "wielkopolskie", "zachodniopomorskie",
    }
    if voivodeship in polish_voivodeships:
        return True
    # Fallback: szukaj polskiego kodu pocztowego w adresie (XX-XXX)
    if re.search(r"\d{2}-\d{3}", address):
        return True
    return False


# =============================================================================
# ETAP 1: OFERENCI Z WYNIKÓW KONKURSÓW POLONIJNYCH MSZ/KPRM (dane w PDF)
# =============================================================================
# Selektory i struktura zweryfikowane przez realne pobranie stron i
# załączników (nie zgadywane). Szczegóły patrz historia rozmowy z 05.07.2026.

def pobierz_zalaczniki_pdf(article_url: str, session: requests.Session, archiwum_wayback: bool) -> list[dict]:
    """
    Pobiera stronę artykułu z wynikami konkursu i zwraca listę załączników PDF.

    Każdy element listy to słownik z kluczami: etykieta, url.
    Selektor a.file-download jest identyczny na stronach MSZ i na
    zarchiwizowanych stronach KPRM w Wayback Machine (ten sam szablon CMS gov.pl).
    """
    soup = fetch_page(article_url, session)
    if soup is None:
        log_event("ZALACZNIKI_ARTYKULU", False, f"Nie można pobrać artykułu: {article_url}")
        return []

    baza_url = "https://web.archive.org" if archiwum_wayback else "https://www.gov.pl"

    zalaczniki = []
    for a in soup.select("a.file-download"):
        href = a.get("href", "")
        if not href:
            continue
        etykieta = a.get("aria-label", "") or a.get_text(strip=True)
        etykieta = re.sub(r"\s*Rozmiar.*$", "", etykieta, flags=re.S).strip()
        etykieta = re.sub(r"^Pobierz plik\s*", "", etykieta).strip()
        if archiwum_wayback:
            # Tryb surowy "id_" — zwykły tryb odtwarzania Wayback Machine
            # potrafi zwrócić błąd 500 dla dużych plików binarnych (PDF).
            href = re.sub(r"^(/web/\d{14})/", r"\1id_/", href)
        url_pelny = href if href.startswith("http") else baza_url + href
        zalaczniki.append({"etykieta": etykieta, "url": url_pelny})

    log_event("ZALACZNIKI_ARTYKULU", True, f"{article_url} — znaleziono {len(zalaczniki)} załączników PDF")
    return zalaczniki


def _naglowek_pasuje(naglowek: str, slowa_kluczowe: list[str]) -> bool:
    """Sprawdza, czy nagłówek kolumny (po usunięciu znaków nowej linii) zawiera któreś ze słów kluczowych."""
    znorm = (naglowek or "").replace("\n", " ").strip().lower()
    return any(slowo in znorm for slowo in slowa_kluczowe)


def _znajdz_indeks_kolumny(naglowki: list[str], slowa_kluczowe: list[str]) -> int | None:
    """Zwraca indeks pierwszej kolumny pasującej do słów kluczowych, albo None."""
    for i, naglowek in enumerate(naglowki):
        if _naglowek_pasuje(naglowek, slowa_kluczowe):
            return i
    return None


def _komorka(wiersz: list, idx: int | None) -> str:
    """Bezpiecznie zwraca komórkę o danym indeksie (część wierszy w PDF bywa krótsza niż nagłówek)."""
    if idx is None or idx >= len(wiersz):
        return ""
    return (wiersz[idx] or "").replace("\n", " ").strip()


def wyciagnij_wiersze_z_pdf(zawartosc_pdf: bytes, zrodlo_opis: str) -> list[dict] | None:
    """
    Wyciąga z pliku PDF listę ofert: Oferent, Nazwa projektu, Kwota.

    Układ kolumn różni się między latami (np. 5 kolumn w 2026 roku,
    12 kolumn w 2023 roku), dlatego kolumny znajdujemy po treści nagłówka,
    a nie po stałej pozycji. Niektóre pliki mają dodatkowo wiersz tytułowy
    NAD właściwym nagłówkiem tabeli, dlatego szukamy wiersza nagłówkowego
    w całym dokumencie po dosłownej treści "Oferent", a nie zakładamy, że
    to zawsze pierwszy wiersz pierwszej strony.
    Jeśli nigdzie nie znajdziemy takiej kolumny, zwraca None i loguje
    zdarzenie — NIE zgadujemy układu.
    """
    wszystkie_wiersze = []
    with pdfplumber.open(BytesIO(zawartosc_pdf)) as pdf:
        for strona in pdf.pages:
            tabela = strona.extract_table()
            if tabela:
                wszystkie_wiersze.extend(tabela)

    if not wszystkie_wiersze:
        log_event("PARSOWANIE_PDF", False, f"{zrodlo_opis} — nie wykryto żadnej tabeli")
        return None

    idx_naglowka = next(
        (i for i, w in enumerate(wszystkie_wiersze) if _znajdz_indeks_kolumny(w, ["oferent"]) is not None),
        None,
    )
    if idx_naglowka is None:
        log_event("PARSOWANIE_PDF", False, f"{zrodlo_opis} — nie znaleziono kolumny 'Oferent'")
        return None

    naglowki = wszystkie_wiersze[idx_naglowka]
    dane = wszystkie_wiersze[idx_naglowka + 1:]

    idx_oferent = _znajdz_indeks_kolumny(naglowki, ["oferent"])
    idx_projekt = _znajdz_indeks_kolumny(naglowki, ["nazwa projektu", "tytuł oferty", "tytul oferty"])

    # Niektóre lata mają osobną kolumnę kwoty przyznanej na każdy rok
    # realizacji projektu (np. 2023: kolumny na 2023, 2024 i 2025 rok).
    # Bierzemy WSZYSTKIE takie kolumny, nie tylko pierwszą — inaczej
    # projekt sfinansowany dopiero w kolejnym roku wyglądałby jak odrzucony.
    idx_kwota_lista = [
        i for i, h in enumerate(naglowki) if _naglowek_pasuje(h, ["proponowana", "dotacj"])
    ]
    if not idx_kwota_lista:
        idx_kwota_lista = [i for i, h in enumerate(naglowki) if _naglowek_pasuje(h, ["kwota"])]

    wiersze = []
    for wiersz in dane:
        if not wiersz or not any(wiersz):
            continue
        # Pomiń ewentualne powtórzenie wiersza nagłówkowego na kolejnej stronie.
        if _komorka(wiersz, idx_oferent).lower() == "oferent":
            continue

        oferent = _komorka(wiersz, idx_oferent)
        if not oferent:
            continue
        projekt = _komorka(wiersz, idx_projekt) or "nie ustalono"

        czesci_kwoty = []
        for idx in idx_kwota_lista:
            wartosc = _komorka(wiersz, idx)
            if wartosc:
                etykieta_kolumny = (naglowki[idx] or "").replace("\n", " ").strip() if idx < len(naglowki) else ""
                czesci_kwoty.append(
                    f"{etykieta_kolumny}: {wartosc}" if len(idx_kwota_lista) > 1 else wartosc
                )
        kwota = "; ".join(czesci_kwoty) if czesci_kwoty else "nie ustalono"

        wiersze.append({
            "oferent": oferent,
            "projekt": projekt or "nie ustalono",
            "kwota": kwota or "nie ustalono",
        })

    return wiersze


def zbierz_oferentow_z_konkursow(session: requests.Session, progress: dict):
    """
    Etap 1: dla każdego skonfigurowanego artykułu z wynikami konkursu polonijnego
    pobiera załączniki PDF (pomijając zbiorcze listy "wszystkich zgłoszonych"),
    wyciąga z nich oferentów — zarówno tych, którym przyznano dotację, jak i
    odrzuconych/nieprzyjętych do dofinansowania — i zapisuje do pliku pośredniego.
    """
    seed_wb, seed_ws, seed_row = create_seed_workbook()
    przetworzone = set(progress.get("przetworzone_zalaczniki", []))

    for konkurs in CONFIG["konkursy_polonijne"]:
        zalaczniki = pobierz_zalaczniki_pdf(konkurs["url"], session, konkurs["archiwum_wayback"])
        polite_delay()

        zalaczniki_do_wziecia = [
            z for z in zalaczniki
            if not any(fraza in z["etykieta"].lower() for fraza in CONFIG["zalaczniki_pomijane_frazy"])
        ]

        for zalacznik in tqdm(zalaczniki_do_wziecia, desc=f"[{konkurs['rok']}] {konkurs['nazwa_konkursu'][:30]}", unit="zal"):
            if zalacznik["url"] in przetworzone:
                continue

            try:
                resp = session.get(zalacznik["url"], headers=HEADERS, timeout=CONFIG["request_timeout"])
                resp.raise_for_status()
            except requests.RequestException as e:
                log_event("POBIERANIE_ZALACZNIKA", False, f"{zalacznik['url']} — {e}")
                przetworzone.add(zalacznik["url"])
                progress["przetworzone_zalaczniki"] = list(przetworzone)
                save_progress(progress)
                polite_delay()
                continue

            wiersze = wyciagnij_wiersze_z_pdf(resp.content, zalacznik["url"])

            if wiersze is None:
                log_event(
                    "STRUKTURA_NIEROZPOZNANA", False,
                    f"{zalacznik['url']} ({zalacznik['etykieta']}) — wymaga ręcznej weryfikacji, pominięto",
                )
            else:
                for wiersz in wiersze:
                    seed_ws.cell(row=seed_row, column=1, value=konkurs["rok"])
                    seed_ws.cell(row=seed_row, column=2, value=konkurs["nazwa_konkursu"])
                    seed_ws.cell(row=seed_row, column=3, value=zalacznik["etykieta"])
                    seed_ws.cell(row=seed_row, column=4, value=wiersz["oferent"])
                    seed_ws.cell(row=seed_row, column=5, value=wiersz["projekt"])
                    seed_ws.cell(row=seed_row, column=6, value=wiersz["kwota"])
                    seed_ws.cell(row=seed_row, column=7, value=konkurs["url"])
                    seed_ws.cell(row=seed_row, column=8, value=zalacznik["url"])
                    seed_row += 1
                    seed_wb.save(CONFIG["seed_file"])
                    progress["total_scraped"] += 1
                    save_progress(progress)

                log_event("ZALACZNIK_PRZETWORZONY", True, f"{zalacznik['url']} — {len(wiersze)} ofert")

            przetworzone.add(zalacznik["url"])
            progress["przetworzone_zalaczniki"] = list(przetworzone)
            save_progress(progress)
            polite_delay()

    seed_wb.save(CONFIG["seed_file"])
    log_event("KONIEC_ETAPU_1", True, f"Plik pośredni: {CONFIG['seed_file']}")


def create_seed_workbook() -> tuple:
    """Tworzy nowy plik pośredni etapu 1 albo wczytuje istniejący (dopisywanie, nie nadpisywanie)."""
    seed_path = Path(CONFIG["seed_file"])
    seed_path.parent.mkdir(parents=True, exist_ok=True)

    if seed_path.exists():
        wb = load_workbook(seed_path)
        ws = wb.active
        next_row = ws.max_row + 1
        log_event("WCZYTANIE_PLIKU_POSREDNIEGO", True, f"Kontynuacja od wiersza {next_row}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Oferenci — konkursy polonijne"
        for col_idx, col_name in enumerate(SEED_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, name="Arial")
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        next_row = 2
        log_event("UTWORZENIE_PLIKU_POSREDNIEGO", True, CONFIG["seed_file"])

    return wb, ws, next_row


# =============================================================================
# ETAP 3: WZBOGACANIE PRZEZ OFICJALNE API KRS I BUDOWA FINALNEJ BAZY
# =============================================================================
# Struktura odpowiedzi API zweryfikowana przez realne zapytania podczas
# ręcznej weryfikacji Etapu 2 (patrz historia rozmowy) — pola danePodmiotu
# (nazwa, formaPrawna, identyfikatory.nip) oraz siedzibaIAdres (siedziba.
# wojewodztwo, adres.ulica/nrDomu/nrLokalu/kodPocztowy/miejscowosc/poczta).
# API nie zwraca numeru REGON ani danych kontaktowych (telefon, e-mail,
# WWW) — to wymagałoby osobno API REGON (GUS BIR1.1, potrzebny klucz
# produkcyjny) i osobnego etapu scrapowania stron WWW organizacji.

PREFIKSY_ULIC = {"UL": "ul.", "AL": "al.", "PL": "pl.", "OS": "os.", "PLAC": "plac"}


def _tytul(tekst: str) -> str:
    """Zamienia zapis WIELKIMI LITERAMI z API na zapis z wielkiej litery każdego słowa."""
    return tekst.title() if tekst else ""


def _sformatuj_ulice(ulica: str) -> str:
    """
    API zwraca ulicę czasem z prefiksem (UL./AL./PL.), czasem bez.
    Prefiks zapisujemy małą literą zgodnie z przykładem z formatki
    ("ul. Mickiewicza 12"), resztę nazwy z wielkiej litery każdego słowa.
    """
    czesci = ulica.strip().split(None, 1)
    if len(czesci) == 2 and czesci[0].rstrip(".").upper() in PREFIKSY_ULIC:
        prefiks = PREFIKSY_ULIC[czesci[0].rstrip(".").upper()]
        return f"{prefiks} {_tytul(czesci[1])}"
    return _tytul(ulica)


def zbuduj_adres_korespondencyjny(adres: dict) -> str:
    """
    Buduje adres korespondencyjny z pól API KRS.
    Adresy wiejskie bez nazwy ulicy (np. "Łukomie 62") obsługiwane osobno —
    numer domu wtedy stoi bezpośrednio przy nazwie miejscowości.
    """
    if not adres:
        return "brak"

    ulica = adres.get("ulica")
    nr_domu = adres.get("nrDomu") or ""
    nr_lokalu = adres.get("nrLokalu") or ""
    kod = adres.get("kodPocztowy") or ""
    miejscowosc = adres.get("miejscowosc") or ""
    poczta = adres.get("poczta") or miejscowosc

    numer = f"{nr_domu}/{nr_lokalu}" if (nr_domu and nr_lokalu) else (nr_domu or nr_lokalu)

    if ulica:
        pierwsza_czesc = f"{_sformatuj_ulice(ulica)} {numer}".strip()
    else:
        pierwsza_czesc = f"{_tytul(miejscowosc)} {numer}".strip()

    druga_czesc = f"{kod} {_tytul(poczta)}".strip()

    czesci = [c for c in (pierwsza_czesc, druga_czesc) if c]
    return ", ".join(czesci) if czesci else "brak"


def wczytaj_cache_krs() -> dict:
    """Wczytuje cache odpowiedzi API KRS (wznowienie po awarii bez ponownych zapytań)."""
    p = Path(CONFIG["cache_dane_krs_file"])
    if p.exists():
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}


def zapisz_cache_krs(cache: dict):
    """Zapisuje cache po KAŻDYM nowym zapytaniu do API — zgodnie z polityką wznowień."""
    p = Path(CONFIG["cache_dane_krs_file"])
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def _pobierz_odpis_krs(numer_krs: str, rejestr: str) -> dict | None:
    """
    Jedno zapytanie do API KRS z jedną powtórką po błędzie sieci (5 s).
    Zwraca None, gdy podmiotu nie ma w tym rejestrze (HTTP 204) albo
    gdy obie próby zawiodły.
    """
    url = CONFIG["api_krs_url_wzor"].format(krs=numer_krs, rejestr=rejestr)
    for probka in (1, 2):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=CONFIG["request_timeout"])
        except requests.RequestException as e:
            log_event(f"API_KRS {numer_krs}", False, f"próba {probka} — {e}")
            if probka == 1:
                time.sleep(CONFIG["retry_delay"])
            continue

        if resp.status_code == 204:
            return None
        try:
            resp.raise_for_status()
        except requests.RequestException as e:
            log_event(f"API_KRS {numer_krs}", False, f"próba {probka} — {e}")
            if probka == 1:
                time.sleep(CONFIG["retry_delay"])
            continue

        return resp.json()

    return None


def pobierz_dane_podmiotu_z_krs(numer_krs: str, cache: dict) -> dict | None:
    """
    Pobiera z oficjalnego API KRS aktualną nazwę, formę prawną, NIP,
    województwo i adres siedziby danego numeru KRS. Najpierw próbuje
    rejestru stowarzyszeń/fundacji (S), a jeśli brak wpisu — rejestru
    przedsiębiorców (P), np. dla spółek z o.o. Wynik cache'owany.
    """
    if numer_krs in cache:
        return cache[numer_krs]

    dane = _pobierz_odpis_krs(numer_krs, "S")
    rejestr_uzyty = "S"
    if dane is None:
        dane = _pobierz_odpis_krs(numer_krs, "P")
        rejestr_uzyty = "P"

    if dane is None:
        log_event("API_KRS_BRAK_WPISU", False, numer_krs)
        cache[numer_krs] = None
        zapisz_cache_krs(cache)
        return None

    dzial1 = dane.get("odpis", {}).get("dane", {}).get("dzial1", {})
    dane_podmiotu = dzial1.get("danePodmiotu") or {}
    siedziba_i_adres = dzial1.get("siedzibaIAdres") or {}

    wynik = {
        "nazwa": dane_podmiotu.get("nazwa") or "",
        "forma_prawna": dane_podmiotu.get("formaPrawna") or "",
        "nip": (dane_podmiotu.get("identyfikatory") or {}).get("nip") or "",
        "wojewodztwo": (siedziba_i_adres.get("siedziba") or {}).get("wojewodztwo") or "",
        "adres": siedziba_i_adres.get("adres") or {},
        "rejestr": rejestr_uzyty,
    }
    cache[numer_krs] = wynik
    zapisz_cache_krs(cache)
    log_event("API_KRS", True, f"{numer_krs} ({rejestr_uzyty})")
    return wynik


def wczytaj_mape_krs() -> dict:
    """
    Wczytuje mapę nazwa-oferenta → numer-KRS z Etapu 2 i buduje słownik
    dopasowań: dokładny zapis nazwy tak, jak wystąpił w pliku pośrednim
    Etapu 1 → LISTA rekordów mapy (zwykle jednoelementowa; dwuelementowa
    tylko dla jednego znanego wiersza, w którym dwie organizacje figurują
    we wspólnej komórce — patrz "FUNDACJA CINEMA MEDIA ART, FINDACJA
    MEDIA WORLD CUP" w mapie).
    """
    with open(CONFIG["mapa_krs_file"], encoding="utf-8") as f:
        rekordy = json.load(f)

    dopasowania = defaultdict(list)
    for rekord in rekordy:
        klucze = set(rekord.get("warianty_zapisu", [])) | {rekord["nazwa_znormalizowana"]}
        for klucz in klucze:
            dopasowania[klucz].append(rekord)

    return dopasowania


def zbuduj_charakterystyke(rekord: dict, lata: set) -> str:
    """
    Buduje krótką charakterystykę (1–3 zdania) wyłącznie z faktów znanych
    z konkursów polonijnych — KRS nie podaje dziedziny działalności, więc
    jej NIE zgadujemy (kolumna "Branża / Typ" zostaje "brak").
    """
    lata_tekst = ", ".join(sorted(lata))
    zdanie = f"Oferent w konkursach polonijnych MSZ/KPRM w latach: {lata_tekst}."

    opisy_statusu = {
        "nie_znaleziono_wyszukiwaniem": "Nie znaleziono numeru KRS — wymaga ręcznej weryfikacji.",
        "niejednoznaczne_kilka_pasujacych_podmiotow": "Kilka podmiotów o tej samej nazwie w KRS — wymaga ręcznego wskazania właściwego.",
        "nie_dotyczy_brak_w_KRS": "Podmiot nie figuruje w KRS (inna forma prawna albo wykreślony z rejestru).",
    }
    opis_statusu = opisy_statusu.get(rekord["status"], "")
    if opis_statusu:
        zdanie += " " + opis_statusu

    return zdanie


def zbuduj_wzbogacona_baze(progress: dict):
    """
    Etap 3: dla każdego oferenta z pliku pośredniego Etapu 1 o potwierdzonym
    numerze KRS (Etap 2) pobiera z oficjalnego API KRS dane rejestrowe
    i zapisuje JEDEN wiersz na organizację (deduplikacja po KRS) w finalnym
    pliku wg kolumn formatki (OUTPUT_COLUMNS). Organizacje bez potwierdzonego
    numeru KRS też trafiają do wyniku — z polami rejestrowymi "brak" i
    wyjaśnieniem w charakterystyce — żeby żaden realny podmiot nie zniknął.
    Tylko artefakty tabeli PDF (status nie_dotyczy_artefakt_tabeli) są
    pomijane całkowicie, bo to nie są prawdziwe organizacje.
    """
    dopasowania = wczytaj_mape_krs()
    cache_krs = wczytaj_cache_krs()

    seed_wb = load_workbook(CONFIG["seed_file"])
    seed_ws = seed_wb.active
    naglowki_seed = [c.value for c in next(seed_ws.iter_rows(min_row=1, max_row=1))]
    idx_oferent = naglowki_seed.index("Oferent")
    idx_rok = naglowki_seed.index("Rok")
    idx_url_artykulu = naglowki_seed.index("URL artykułu")

    grupy: dict[str, dict] = {}
    niedopasowane = set()

    for wiersz in seed_ws.iter_rows(min_row=2, values_only=True):
        oferent_surowy = wiersz[idx_oferent]
        if not oferent_surowy:
            continue

        pasujace_rekordy = dopasowania.get(str(oferent_surowy))
        if not pasujace_rekordy:
            niedopasowane.add(str(oferent_surowy))
            continue

        for rekord in pasujace_rekordy:
            if rekord["status"] == "nie_dotyczy_artefakt_tabeli":
                continue

            klucz_grupy = rekord.get("krs") or f"BEZ_KRS::{rekord['nazwa_znormalizowana']}"
            grupa = grupy.setdefault(klucz_grupy, {"rekord_mapy": rekord, "lata": set(), "urle": set()})
            grupa["lata"].add(str(wiersz[idx_rok]))
            grupa["urle"].add(str(wiersz[idx_url_artykulu]))

    if niedopasowane:
        log_event(
            "OFERENCI_NIEDOPASOWANI", False,
            f"{len(niedopasowane)} nazw z pliku pośredniego nie znaleziono w mapie KRS: "
            + "; ".join(sorted(niedopasowane)[:20]),
        )

    out_wb, out_ws, wiersz_wyjsciowy = create_output_workbook()
    juz_zapisane = set(progress.get("etap3_zapisane_grupy", []))
    dzis = datetime.now().strftime(CONFIG["date_format"])

    do_przetworzenia = [(k, g) for k, g in sorted(grupy.items()) if k not in juz_zapisane]
    log.info(f"Etap 3: {len(do_przetworzenia)}/{len(grupy)} organizacji do przetworzenia (reszta już zapisana)")

    for klucz_grupy, grupa in tqdm(do_przetworzenia, desc="Budowanie bazy", unit="org"):
        rekord = grupa["rekord_mapy"]
        numer_krs = rekord.get("krs")

        dane_api = pobierz_dane_podmiotu_z_krs(numer_krs, cache_krs) if numer_krs else None
        if numer_krs:
            polite_delay()

        if dane_api:
            # Nazwa zostaje WIELKIMI LITERAMI — tak jak jest zapisana w odpisie
            # KRS (to jest konwencja samego rejestru, nie usterka danych;
            # "zgodna z rejestrem" z procedury oznacza dosłowną zgodność).
            nazwa = dane_api["nazwa"] or rekord["nazwa_znormalizowana"]
            wojewodztwo = dane_api["wojewodztwo"].lower() if dane_api["wojewodztwo"] else "brak"
            adres_tekst = zbuduj_adres_korespondencyjny(dane_api["adres"])
            nip = dane_api["nip"] or "brak"
        else:
            nazwa = rekord["nazwa_znormalizowana"]
            wojewodztwo = "brak"
            adres_tekst = "brak"
            nip = "brak"

        rekord_wyjsciowy = {
            "Kategoria": "organizacja pozarządowa",
            "Branża / Typ": "brak",
            "Nazwa": nazwa,
            "Adres korespondencyjny": adres_tekst,
            "Województwo": wojewodztwo,
            "Numer telefonu": "brak",
            "Adres e-mail": "brak",
            "Strona WWW": "brak",
            "Profil w mediach społecznościowych": "brak",
            "Osoba kontaktowa": "nie ustalono",
            "Numer telefonu do osoby kontaktowej": "brak",
            "Adres e-mail do osoby kontaktowej": "brak",
            "Data pozyskania informacji": dzis,
            "Krótka charakterystyka podmiotu": zbuduj_charakterystyke(rekord, grupa["lata"]),
            "URL źródła": "; ".join(sorted(grupa["urle"])),
            "KRS": numer_krs or "brak",
            "REGON": "brak",
            "NIP": nip,
        }

        write_record(out_ws, wiersz_wyjsciowy, rekord_wyjsciowy)
        wiersz_wyjsciowy += 1
        out_wb.save(CONFIG["output_file"])

        juz_zapisane.add(klucz_grupy)
        progress["etap3_zapisane_grupy"] = list(juz_zapisane)
        save_progress(progress)

        log_event("ORGANIZACJA_ZAPISANA", True, f"{nazwa[:50]} (KRS {numer_krs or 'brak'})")

    log_event("KONIEC_ETAPU_3", True, f"Łącznie {len(juz_zapisane)} organizacji w {CONFIG['output_file']}")


# =============================================================================
# ETAP 4: DANE KONTAKTOWE ZE STRON WWW ORGANIZACJI
# =============================================================================
# Adres strony każdej organizacji znajdujemy ręcznie (wyszukiwanie po nazwie
# z rejestru, nie po numerze KRS — numer KRS prawie nigdy nie występuje na
# własnej stronie organizacji, za to bardzo często na stronach zbiorczych
# typu rejestr.io czy aleo.com, których chcemy unikać). To, co JEST
# zautomatyzowane poniżej, to pobranie znanego adresu i wyciągnięcie z niego
# telefonu, e-maila i profilu social media uniwersalnymi wzorcami tekstowymi —
# to działa niezależnie od struktury konkretnej strony, więc nie jest
# "zgadywaniem selektorów CSS" zakazanym w CLAUDE.md.

WZORZEC_EMAIL = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")

# Domeny, które ewidentnie NIE są własną stroną organizacji, tylko
# rejestrem zbiorczym — pomijamy je przy wyborze adresu WWW.
DOMENY_REJESTROW_DO_POMINIECIA = [
    "rejestr.io", "aleo.com", "krs-pobierz.pl", "krs-online.com.pl",
    "ngo.pl", "imsig.pl", "infoveriti.pl", "gowork.pl", "panoramafirm.pl",
    "bizraport.pl", "pitax.pl", "moc-danych.pl", "wyszukiwarkakrs.pl",
    "dnb.com", "emis.com",
    # Mapy, katalogi firm, wyszukiwarki i wideo — nigdy nie są własną
    # stroną organizacji, choć bywają wysoko w wynikach wyszukiwania.
    "yandex.", "google.com/maps", "maps.google", "mapy.cz", "targeo.pl",
    "bing.com", "duckduckgo.com", "wikipedia.org", "youtube.com",
    "facebook.com", "instagram.com", "linkedin.com", "twitter.com", "x.com",
    "cylex-polska.pl", "cylex.pl", "firmy.net", "biznes-firmy.pl",
    "oferteo.pl", "panoramafirm.eu", "pkt.pl", "firmy.pwginfo.pl",
    "mapa.targeo.pl", "zumi.pl", "firmania.pl", "godzinyotwarcia24.pl",
    "findglocal.com", "giveradar.com", "onas.org.pl",
]


def czy_adres_to_rejestr(url: str) -> bool:
    """Sprawdza, czy podany adres to strona zbiorczego rejestru, a nie własna strona organizacji."""
    return any(domena in url.lower() for domena in DOMENY_REJESTROW_DO_POMINIECIA)


def _wyciagnij_emaile(tekst: str) -> list[str]:
    """Zwraca unikalne adresy e-mail znalezione w tekście, w kolejności występowania."""
    znalezione = []
    for dopasowanie in WZORZEC_EMAIL.findall(tekst):
        if dopasowanie not in znalezione:
            znalezione.append(dopasowanie)
    return znalezione


# Kolejność priorytetu adresów e-mail organizacji: sekretariat > biuro > reszta
# (zgodnie z zasadami CLAUDE.md i docs/zrodla_polonia.md).
PRIORYTET_EMAIL = ["sekretariat", "biuro", "kontakt", "info", "office"]


def _posortuj_wg_priorytetu(emaile: list[str]) -> list[str]:
    """Sortuje e-maile tak, by sekretariat/biuro/kontakt/info trafiały na początek listy."""
    def klucz(adres):
        adres_maly = adres.lower()
        for i, slowo in enumerate(PRIORYTET_EMAIL):
            if slowo in adres_maly:
                return i
        return len(PRIORYTET_EMAIL)
    return sorted(emaile, key=klucz)


def _wyciagnij_telefony(tekst: str) -> list[str]:
    """
    Zwraca unikalne numery telefonu znalezione w tekście, wyszukane i
    sformatowane biblioteką phonenumbers (krajowy format polski). Biblioteka
    waliduje numer wg rzeczywistego planu numeracji (poprawny kierunkowy,
    poprawna długość), dzięki czemu odrzuca przypadkowe sekwencje cyfr typu
    NIP, REGON czy numer KRS znacznie skuteczniej niż sam regex.
    """
    znalezione = []
    for dopasowanie in phonenumbers.PhoneNumberMatcher(tekst, "PL"):
        numer = phonenumbers.format_number(dopasowanie.number, phonenumbers.PhoneNumberFormat.NATIONAL)
        if numer not in znalezione:
            znalezione.append(numer)
    return znalezione


def _wyciagnij_social_linki(parser: HTMLParser) -> dict:
    """Zwraca pierwszy znaleziony link do każdej platformy social media, wg kolejności priorytetu w CONFIG."""
    linki = [(a.attributes.get("href") or "") for a in parser.css("a[href]")]
    wynik = {}
    for domena in CONFIG["domeny_social_media"]:
        for link in linki:
            if domena in link.lower():
                wynik[domena] = link
                break
    return wynik


def _tekst_i_parser_ze_strony(url: str, client: httpx.Client) -> tuple[str, HTMLParser] | None:
    """Pobiera stronę przez httpx (podstawowe narzędzie, patrz CLAUDE.md) i zwraca (czysty tekst, HTMLParser)."""
    try:
        resp = client.get(url, headers=HEADERS, timeout=CONFIG["request_timeout"], follow_redirects=True)
        resp.raise_for_status()
    except httpx.HTTPError as e:
        log_event(f"POBIERANIE_KONTAKTU {url[:60]}", False, str(e))
        return None
    parser = HTMLParser(resp.text)
    return parser.text(separator=" ", deep=True, strip=True), parser


def _tekst_i_html_playwright(url: str) -> tuple[str, str] | None:
    """
    Pobiera stronę renderowaną przez JavaScript przez Playwright (Chromium
    headless) — zapasowe narzędzie, używane tylko gdy httpx zwróci
    podejrzanie mało tekstu (patrz prog_dlugosci_tekstu_js w CONFIG).
    """
    from playwright.sync_api import sync_playwright

    # Środowisko deweloperskie, w którym ten kod był pisany, ma preinstalowaną
    # przeglądarkę pod tą ścieżką (patrz zmienna PLAYWRIGHT_BROWSERS_PATH) —
    # używamy jej wtedy wprost, zamiast domyślnej (czasem niedopasowanej do
    # wersji pakietu playwright). Na docelowym komputerze bez tej ścieżki
    # Playwright samodzielnie zarządza własną, dopasowaną przeglądarką.
    wbudowana_przegladarka = Path("/opt/pw-browsers/chromium")
    argumenty_launch = {"headless": True}
    if wbudowana_przegladarka.exists():
        argumenty_launch["executable_path"] = str(wbudowana_przegladarka)

    try:
        with sync_playwright() as p:
            przegladarka = p.chromium.launch(**argumenty_launch)
            try:
                strona = przegladarka.new_page(user_agent=HEADERS["User-Agent"])
                strona.set_default_timeout(CONFIG["request_timeout"] * 1000)
                strona.goto(url, wait_until="domcontentloaded")
                tekst = strona.inner_text("body")
                html = strona.content()
            finally:
                przegladarka.close()
    except Exception as e:
        log_event(f"PLAYWRIGHT {url[:60]}", False, str(e))
        return None
    return tekst, html


def _pobierz_jedna_podstrone(adres: str, client: httpx.Client) -> tuple[str, HTMLParser, str] | None:
    """
    Pobiera jedną stronę httpx (podstawowe narzędzie — lekkie i szybkie).
    Jeśli zwrócony tekst jest podejrzanie krótki — prawdopodobnie strona
    jest renderowana przez JavaScript — sięga po Playwright jako zapasowe
    narzędzie (patrz CLAUDE.md, sekcja metodyki pozyskiwania danych).
    Zwraca (tekst, parser, metoda) albo None.
    """
    pobrane_httpx = _tekst_i_parser_ze_strony(adres, client)
    if pobrane_httpx is not None:
        tekst, parser = pobrane_httpx
        if len(tekst) >= CONFIG["prog_dlugosci_tekstu_js"]:
            return tekst, parser, "httpx"

    pobrane_playwright = _tekst_i_html_playwright(adres)
    if pobrane_playwright is not None:
        tekst, html = pobrane_playwright
        return tekst, HTMLParser(html), "playwright"

    if pobrane_httpx is not None:
        # httpx zwróciło niewiele tekstu, ale Playwright zawiódł —
        # lepsze to niż nic.
        tekst, parser = pobrane_httpx
        return tekst, parser, "httpx (mało tekstu)"

    return None


def zbierz_dane_kontaktowe_z_adresu(url: str) -> dict:
    """
    Pobiera stronę główną i typowe podstrony kontaktowe (kontakt/o nas/contact),
    wyciąga z nich e-maile, telefony i linki do mediów społecznościowych.
    httpx + selectolax to podstawowe narzędzie (szybkie, lekkie), Playwright
    to zapasowe narzędzie dla stron renderowanych przez JavaScript.
    """
    wynik = {"emaile": [], "telefony": [], "social": {}, "url_zrodlowy": url, "metoda": "brak"}

    # Podstrony kontaktowe dobudowujemy od głównej domeny (schemat + host),
    # NIE od dokładnej ścieżki podanego URL-a — inaczej dla adresu typu
    # https://przyklad.pl/kontakt/ powstałaby błędna ścieżka
    # https://przyklad.pl/kontakt/kontakty.
    czesci_url = urllib.parse.urlsplit(url)
    baza = f"{czesci_url.scheme}://{czesci_url.netloc}"

    adresy_do_sprawdzenia = [url]
    for podstrona in CONFIG["podstrony_kontaktowe"]:
        adres_podstrony = f"{baza}/{podstrona}"
        if adres_podstrony not in adresy_do_sprawdzenia:
            adresy_do_sprawdzenia.append(adres_podstrony)

    metody_uzyte = set()
    with httpx.Client() as client:
        for adres in adresy_do_sprawdzenia:
            pobrane = _pobierz_jedna_podstrone(adres, client)
            if pobrane is None:
                continue
            tekst, parser, metoda = pobrane
            metody_uzyte.add(metoda)

            for email in _wyciagnij_emaile(tekst):
                if email not in wynik["emaile"]:
                    wynik["emaile"].append(email)
            for telefon in _wyciagnij_telefony(tekst):
                if telefon not in wynik["telefony"]:
                    wynik["telefony"].append(telefon)
            for domena, link in _wyciagnij_social_linki(parser).items():
                wynik["social"].setdefault(domena, link)

            polite_delay()

            if wynik["emaile"] and wynik["telefony"]:
                wynik["url_zrodlowy"] = adres
                break

    wynik["metoda"] = " + ".join(sorted(metody_uzyte)) if metody_uzyte else "brak"
    return wynik


# Flaga globalna — jeśli ddgs raz zwróci sygnał blokady/CAPTCHA, wyłączamy
# automatyczne wyszukiwanie na resztę uruchomienia skryptu i wymuszamy
# ręczne wyszukiwanie adresu WWW (patrz CLAUDE.md — nie omijamy zabezpieczeń
# antybotowych, nawet pośrednio przez ponawianie prób).
_ddgs_zablokowane = False


def znajdz_strone_organizacji(nazwa: str) -> str | None:
    """
    Szuka oficjalnej strony organizacji przez bibliotekę ddgs, pomijając
    strony zbiorczych rejestrów (patrz DOMENY_REJESTROW_DO_POMINIECIA).
    Jeśli DuckDuckGo zwróci sygnał blokady/limitu zapytań — CO JEST INNE
    NIŻ zwykły brak wyników — wyłącza automatyczne wyszukiwanie do końca
    uruchomienia skryptu i zwraca None, żeby operator wyszukał stronę ręcznie.
    """
    global _ddgs_zablokowane
    if _ddgs_zablokowane:
        return None

    from ddgs import DDGS
    from ddgs.exceptions import DDGSException

    try:
        wyniki = list(DDGS().text(f'"{nazwa}" kontakt', max_results=8))
    except DDGSException as e:
        log_event(
            "DDGS_ZABLOKOWANE", False,
            f"Wyłączam automatyczne wyszukiwanie do końca uruchomienia — wymagane ręczne wyszukiwanie: {e}",
        )
        _ddgs_zablokowane = True
        return None

    for wynik in wyniki:
        adres = wynik.get("href", "")
        if not adres or czy_adres_to_rejestr(adres):
            continue
        czesci = urllib.parse.urlsplit(adres)
        if czesci.scheme not in ("http", "https") or not czesci.netloc:
            continue  # np. względny link śledzący wyszukiwarki, nie prawdziwy URL
        return adres
    return None


def sprawdz_kontakt_organizacji(numer_krs: str, url: str | None = None):
    """
    CLI: dla danego numeru KRS znajduje (jeśli url nie podano — przez ddgs,
    z ręcznym wyszukiwaniem jako zapasową metodą, patrz znajdz_strone_organizacji)
    i pobiera stronę WWW organizacji, wyciąga z niej dane kontaktowe i
    zapisuje je do output/dane_kontaktowe.json pod danym numerem KRS.
    """
    path = Path(CONFIG["dane_kontaktowe_file"])
    with open(path, encoding="utf-8") as f:
        dane = json.load(f)

    rekord = next((r for r in dane if r["krs"] == numer_krs), None)
    if rekord is None:
        log.error(f"Nie znaleziono rekordu KRS {numer_krs} w {path}")
        return

    if url is None:
        url = znajdz_strone_organizacji(rekord["nazwa"])
        if url is None:
            log.warning(f"Nie znaleziono strony dla {rekord['nazwa']} (KRS {numer_krs}) — wyszukaj ręcznie")
            return

    wynik = zbierz_dane_kontaktowe_z_adresu(url)

    emaile_posortowane = _posortuj_wg_priorytetu(wynik["emaile"])

    rekord["strona_www"] = url
    rekord["telefon_ogolny"] = wynik["telefony"][0] if wynik["telefony"] else "brak"
    rekord["email_ogolny"] = emaile_posortowane[0] if emaile_posortowane else "brak"
    rekord["profil_social"] = next(iter(wynik["social"].values()), "brak")
    rekord["zrodlo_url"] = wynik["url_zrodlowy"]
    rekord["status"] = "znaleziono" if (wynik["telefony"] or wynik["emaile"]) else "znaleziono_czesciowo"
    rekord.setdefault("osoba_kontaktowa", "nie ustalono")
    rekord.setdefault("telefon_osoby_kontaktowej", "brak")
    rekord.setdefault("email_osoby_kontaktowej", "brak")
    rekord.setdefault("charakterystyka_realna", None)
    rekord.setdefault("branza_typ", None)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(dane, f, ensure_ascii=False, indent=2)

    print(f"KRS {numer_krs} ({wynik['metoda']}):")
    print(f"  strona: {url}")
    print(f"  e-maile znalezione: {wynik['emaile']}")
    print(f"  telefony znalezione: {wynik['telefony']}")
    print(f"  social: {wynik['social']}")
    log_event("KONTAKT_SPRAWDZONY", True, f"{numer_krs} — {url}")


# =============================================================================
# ZAPIS DO XLSX
# =============================================================================

def create_output_workbook() -> tuple:
    """
    Tworzy nowy plik XLSX z nagłówkami lub wczytuje istniejący.
    Zwraca (workbook, worksheet, następny_wiersz).
    """
    output_path = Path(CONFIG["output_file"])
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.exists():
        wb = load_workbook(output_path)
        ws = wb.active
        next_row = ws.max_row + 1
        log_event("WCZYTANIE_PLIKU_WYJSCIOWEGO", True, f"Kontynuacja od wiersza {next_row}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Baza danych"
        _write_header_row(ws)
        next_row = 2
        log_event("UTWORZENIE_PLIKU_WYJSCIOWEGO", True, CONFIG["output_file"])

    return wb, ws, next_row


def _write_header_row(ws):
    """Zapisuje wiersz nagłówkowy ze stylowaniem."""
    # Kolumna Lp. (wypełnia użytkownik)
    ws["A1"] = "Lp."
    # Pozostałe kolumny
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=2):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, name="Arial")
        cell.fill = PatternFill("solid", fgColor="D9E1F2")  # jasny niebieski
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Szerokości kolumn (przybliżone)
    column_widths = {
        "A": 6,   # Lp.
        "B": 20,  # Kategoria
        "C": 20,  # Branża / Typ
        "D": 35,  # Nazwa
        "E": 35,  # Adres
        "F": 18,  # Województwo
        "G": 30,  # Telefon
        "H": 30,  # E-mail
        "I": 40,  # Strona WWW
        "J": 35,  # Social media
        "K": 25,  # Osoba kontaktowa
        "L": 25,  # Tel. os. kontaktowej
        "M": 30,  # E-mail os. kontaktowej
        "N": 20,  # Data
        "O": 40,  # Opis
        "P": 50,  # URL źródła
        "Q": 14,  # KRS
        "R": 14,  # REGON
        "S": 14,  # NIP
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 30


def write_record(ws, row: int, record: dict):
    """Zapisuje jeden rekord do arkusza."""
    ws.cell(row=row, column=1, value="")  # Lp. — puste, wypełni użytkownik
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=2):
        cell = ws.cell(row=row, column=col_idx, value=record.get(col_name, "brak"))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(name="Arial", size=10)

    # Zebry — naprzemienne kolory wierszy dla czytelności
    if row % 2 == 0:
        fill = PatternFill("solid", fgColor="F2F2F2")
        for col in range(1, len(OUTPUT_COLUMNS) + 2):
            ws.cell(row=row, column=col).fill = fill


def create_rejected_workbook() -> tuple:
    """Tworzy lub wczytuje plik odrzuconych rekordów."""
    rejected_path = Path(CONFIG["rejected_file"])
    rejected_path.parent.mkdir(parents=True, exist_ok=True)

    if rejected_path.exists():
        wb = load_workbook(rejected_path)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Odrzucone"
        # Nagłówki
        for col, header in enumerate(["URL", "Powód odrzucenia", "Czas"], start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, name="Arial")
            cell.fill = PatternFill("solid", fgColor="FFB3B3")  # jasny czerwony
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        next_row = 2

    return wb, ws, next_row


def write_rejected(ws, row: int, url: str, reason: str):
    """Zapisuje odrzucony rekord do arkusza odrzuconych."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=row, column=1, value=url)
    ws.cell(row=row, column=2, value=reason)
    ws.cell(row=row, column=3, value=ts)


# =============================================================================
# GŁÓWNA LOGIKA SCRAPOWANIA
# =============================================================================

def scrape_source(
    source: dict,
    session: requests.Session,
    out_ws,
    out_wb,
    rej_ws,
    rej_wb,
    progress: dict,
    current_row: list,   # mutable wrapper żeby unikać nonlocal
    rej_row: list,
):
    """Przetwarza jedno źródło danych od początku do końca."""
    log_event(f"START_ZRODLA_{source['id'].upper()}", True)

    # Pobierz listę URL-i do przetworzenia
    entity_urls = get_entity_urls_from_source(source, session)

    if not entity_urls:
        log_event(f"BRAK_URLS_{source['id'].upper()}", False, "Brak URL-i do przetworzenia")
        return

    # Filtruj już przetworzone (wznowienie po awarii)
    remaining = [u for u in entity_urls if u not in progress["completed_ids"]]
    log.info(f"Źródło '{source['id']}': {len(remaining)}/{len(entity_urls)} URL-i do przetworzenia")

    for url in tqdm(remaining, desc=f"[{source['id']}]", unit="rek"):
        soup = fetch_page(url, session)

        if soup is None:
            write_rejected(rej_ws, rej_row[0], url, "Błąd pobierania strony — 2 próby nieudane")
            rej_row[0] += 1
            rej_wb.save(CONFIG["rejected_file"])
            log_event("REKORD_ODRZUCONY", False, url)
            progress["completed_ids"].append(url)
            save_progress(progress)
            polite_delay()
            continue

        # Parsuj dane
        record = parse_entity_page(url, soup)

        # Sprawdź czy siedziba jest w Polsce
        if not is_based_in_poland(record):
            write_rejected(rej_ws, rej_row[0], url, "Siedziba poza Polską")
            rej_row[0] += 1
            rej_wb.save(CONFIG["rejected_file"])
            log_event("ODRZUCONO_POZA_PL", False, url)
            progress["completed_ids"].append(url)
            save_progress(progress)
            polite_delay()
            continue

        # Zapisz rekord
        write_record(out_ws, current_row[0], record)
        current_row[0] += 1
        out_wb.save(CONFIG["output_file"])

        # Zapisz postęp
        progress["completed_ids"].append(url)
        progress["total_scraped"] += 1
        save_progress(progress)

        log_event(f"ZAPIS_REKORDU {record.get('Nazwa', url)[:40]}", True)
        polite_delay()

    log_event(f"KONIEC_ZRODLA_{source['id'].upper()}", True, f"Przetworzono {len(remaining)} URL-i")


# =============================================================================
# PUNKT WEJŚCIA
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="Scraper bazy danych organizacji")
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Usuń postęp i zacznij od nowa",
    )
    parser.add_argument(
        "--zbierz-oferentow",
        action="store_true",
        help="Etap 1: zbierz nazwy oferentów z wyników konkursów polonijnych MSZ/KPRM (plik pośredni, nie formatka)",
    )
    parser.add_argument(
        "--zbuduj-baze",
        action="store_true",
        help="Etap 3: wzbogać oferentów danymi z API KRS i zbuduj finalny plik wg formatki",
    )
    parser.add_argument(
        "--sprawdz-kontakt",
        nargs="+",
        metavar="KRS [URL]",
        help="Etap 4: znajdź (przez ddgs, albo podany URL) i pobierz stronę organizacji, wyciągnij telefon/e-mail/social",
    )
    args = parser.parse_args()

    if args.reset:
        reset_progress()
        log.info("Postęp zresetowany. Zaczynamy od początku.")

    log_event("START_SKRYPTU", True, f"Zadanie: {CONFIG['task_name']}")

    # Wczytaj postęp (wznowienie lub nowy start)
    progress = load_progress()
    progress.setdefault("przetworzone_zalaczniki", [])

    if args.zbierz_oferentow:
        session = requests.Session()
        try:
            zbierz_oferentow_z_konkursow(session, progress)
        finally:
            session.close()
        return

    if args.zbuduj_baze:
        zbuduj_wzbogacona_baze(progress)
        return

    if args.sprawdz_kontakt:
        numer_krs = args.sprawdz_kontakt[0]
        url = args.sprawdz_kontakt[1] if len(args.sprawdz_kontakt) > 1 else None
        sprawdz_kontakt_organizacji(numer_krs, url)
        return

    # Inicjalizuj pliki wyjściowe
    out_wb, out_ws, next_out_row = create_output_workbook()
    rej_wb, rej_ws, next_rej_row = create_rejected_workbook()

    # Mutable wrappery dla liczników wierszy
    current_row = [next_out_row]
    rej_row = [next_rej_row]

    # Sesja HTTP (wielokrotne użycie połączenia)
    session = requests.Session()

    # Przetwarzaj źródła według priorytetu
    sources_sorted = sorted(CONFIG["sources"], key=lambda s: s["priority"])

    try:
        for source in sources_sorted:
            scrape_source(
                source, session,
                out_ws, out_wb,
                rej_ws, rej_wb,
                progress,
                current_row, rej_row,
            )
    except KeyboardInterrupt:
        log.info("Przerwano przez użytkownika (Ctrl+C). Postęp zapisany.")
    finally:
        # Ostatni zapis
        out_wb.save(CONFIG["output_file"])
        rej_wb.save(CONFIG["rejected_file"])
        session.close()

    # Podsumowanie
    total_scraped = progress["total_scraped"]
    total_rejected = rej_row[0] - 2  # odejmij nagłówek
    print()
    print("=" * 55)
    print(f"  ✅  Pobrano rekordów:    {total_scraped}")
    print(f"  ❌  Odrzucono rekordów:  {max(0, total_rejected)}")
    print(f"  📄  Plik wyjściowy:      {CONFIG['output_file']}")
    print(f"  📋  Odrzucone:           {CONFIG['rejected_file']}")
    print(f"  📝  Logi:                {CONFIG['log_file']}")
    print("=" * 55)

    log_event("KONIEC_SKRYPTU", True, f"Pobrano: {total_scraped}, Odrzucono: {max(0, total_rejected)}")


if __name__ == "__main__":
    main()
